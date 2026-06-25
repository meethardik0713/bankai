import os
from dotenv import load_dotenv
load_dotenv()
import logging
import openpyxl
import time
import hashlib
import threading
import hmac
import razorpay
from flask import Flask, request, render_template, send_file, abort, session, redirect, jsonify, Response
from flask_cors import CORS
from flask_talisman import Talisman
from io import BytesIO
from collections import defaultdict
from werkzeug.utils import secure_filename
from universal_parser import parse_transactions
from core.verifier import run_accuracy_check
from core.post_validator import validate_and_fix
from core.sqlite_indexer import build_index, drop_index, get_db
from core.chat_engine import verify_data, chat as ai_chat
from core.dashboard import run_dashboard
from core.gstr1 import run_gstr1
from supabase import create_client, Client

SUPABASE_URL = os.environ.get('SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY')
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

RAZORPAY_KEY_ID     = os.environ.get('RAZORPAY_KEY_ID')
RAZORPAY_KEY_SECRET = os.environ.get('RAZORPAY_KEY_SECRET')
rzp_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))

from posthog import Posthog
POSTHOG_API_KEY = os.environ.get('POSTHOG_API_KEY', '')
POSTHOG_HOST    = os.environ.get('POSTHOG_HOST', 'https://us.i.posthog.com')
ph = Posthog(project_api_key=POSTHOG_API_KEY, host=POSTHOG_HOST)

# ── FIXED: was calling itself recursively before ──────────
def ph_track(user_id, event, props=None, properties=None):
    p = properties or props or {}
    try:
        ph.capture(distinct_id=str(user_id or 'anonymous'), event=event, properties=p)
    except Exception:
        pass

# ── Logging ───────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger(__name__)

# ── App setup ─────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(32))
CORS(app, resources={r"/api/*": {"origins": "*"}})

is_prod = os.environ.get('RAILWAY_ENVIRONMENT') is not None
Talisman(app,
    force_https=is_prod,
    strict_transport_security=is_prod,
    strict_transport_security_preload=is_prod,
    session_cookie_secure=is_prod,
    content_security_policy={
        'default-src': "'self'",
        'script-src': ["'self'", "'unsafe-inline'", 'https://cdnjs.cloudflare.com', 'https://challenges.cloudflare.com', 'https://checkout.razorpay.com', 'https://checkout.razorpay.com', 'https://cdn.razorpay.com', 'https://www.googletagmanager.com', 'https://www.google-analytics.com', 'https://us-assets.i.posthog.com'],
        'frame-src': ["'self'", 'https://api.razorpay.com', 'https://checkout.razorpay.com', 'https://cdn.razorpay.com'],
        'connect-src': ["'self'", 'https://api.razorpay.com', 'https://checkout.razorpay.com', 'https://lumberjack.razorpay.com', 'https://cdn.razorpay.com', 'https://us.i.posthog.com'],
        'style-src': ["'self'", "'unsafe-inline'", 'https://fonts.googleapis.com'],
        'font-src': ["'self'", 'https://fonts.gstatic.com'],
        'img-src': ["'self'", 'data:', 'https:'],
        'object-src': "'none'",
    }
)

_POSTHOG_JS = (
    b'<script>'
    b'!function(t,e){var o,n,p,r;e.__SV||(window.posthog=e,e._i=[],e.init=function(i,s,a){'
    b'function g(t,e){var o=e.split(".");2==o.length&&(t=t[o[0]],e=o[1]),t[e]=function(){'
    b't.push([e].concat(Array.prototype.slice.call(arguments,0)))}}'
    b'(p=t.createElement("script")).type="text/javascript",p.crossOrigin="anonymous",p.async=!0,'
    b'p.src=s.api_host.replace(".i.posthog.com","-assets.i.posthog.com")+"/static/array.js",'
    b'(r=t.getElementsByTagName("script")[0]).parentNode.insertBefore(p,r);var u=e;'
    b'for(void 0!==a?u=e[a]=[]:a="posthog",u.people=u.people||[],u.toString=function(t){'
    b'var e="posthog";return"posthog"!==a&&(e+="."+a),t||(e+=" (stub)"),e},'
    b'u.people.toString=function(){return u.toString(1)+" (stub)"},'
    b'o="capture identify alias people.set people.set_once set_config register register_once '
    b'unregister opt_out_capturing has_opted_out_capturing opt_in_capturing reset isFeatureEnabled '
    b'onFeatureFlags getFeatureFlag getFeatureFlagPayload reloadFeatureFlags group '
    b'updateEarlyAccessFeatureEnrollment getEarlyAccessFeatures getActiveMatchingSurveys getSurveys '
    b'onSessionId setPersonPropertiesForFlags".split(" "),n=0;n<o.length;n++)g(u,o[n]);'
    b'e._i.push([i,s,a])},e.__SV=1)}(document,window.posthog||(window.posthog=[]));'
    b'posthog.init("' + POSTHOG_API_KEY.encode() + b'",'
    b'{"api_host":"' + POSTHOG_HOST.encode() + b'","person_profiles":"identified_only"});'
    b'</script>'
)

@app.after_request
def inject_ga(response):
    if response.content_type.startswith('text/html'):
        ga_script = b"""
<script async src="https://www.googletagmanager.com/gtag/js?id=G-E9STGNFHT4"></script>
<script>
  window.dataLayer = window.dataLayer || [];
  function gtag(){dataLayer.push(arguments);}
  gtag('js', new Date());
  gtag('config', 'G-E9STGNFHT4');
</script>"""
        response.data = response.data.replace(b'</head>', ga_script + _POSTHOG_JS + b'</head>')
    return response

UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

CHAT_MESSAGE_LIMIT = 50
CHAT_SESSION_HOURS = 48

OWNER_EMAILS = ['hardik101306@gmail.com']

NEW_STATEMENT_ERROR = 'New bank statement detected. Your ₹49 session is locked to one statement. <a href="/pay" style="color:#C9A84C;text-decoration:underline;font-weight:600;">Click here to purchase a new session</a> to analyze a different PDF.'

PLAN_CONFIG = {
    'Basic':    {'price': 10,  'input_tokens': 25000,   'output_tokens': 5000,  'validity_hours': 24},
    'Standard': {'price': 49,  'input_tokens': 100000,  'output_tokens': 10000, 'validity_hours': 75},
    'Pro':      {'price': 99,  'input_tokens': 300000,  'output_tokens': 15000, 'validity_hours': 125},
    'Elite':    {'price': 499, 'input_tokens': 1500000, 'output_tokens': 20000, 'validity_hours': 200},
}

request_counts = defaultdict(list)
RATE_LIMIT  = 10
RATE_WINDOW = 60

_CACHE      = {}
_CACHE_LOCK = threading.Lock()
_CACHE_TTL  = 1800
PAGE_SIZE   = 100
_PDF_MAGIC  = b'%PDF'

def _cache_get(file_hash: str):
    with _CACHE_LOCK:
        entry = _CACHE.get(file_hash)
        if entry and (time.time() - entry['ts']) < _CACHE_TTL:
            return entry
        if entry:
            del _CACHE[file_hash]
        return None


def _cache_set(file_hash: str, transactions: list, filename: str):
    with _CACHE_LOCK:
        if len(_CACHE) >= 10:
            oldest = min(_CACHE, key=lambda k: _CACHE[k]['ts'])
            del _CACHE[oldest]
        _CACHE[file_hash] = {'transactions': transactions, 'filename': filename, 'ts': time.time()}


def _file_hash(file_storage) -> str:
    chunk = file_storage.stream.read(65536)
    file_storage.stream.seek(0)
    return hashlib.md5(chunk).hexdigest()


def is_rate_limited(ip: str) -> bool:
    now = time.time()
    request_counts[ip] = [t for t in request_counts[ip] if now - t < RATE_WINDOW]
    if len(request_counts[ip]) >= RATE_LIMIT:
        return True
    request_counts[ip].append(now)
    return False


def _is_valid_pdf(file_storage) -> bool:
    header = file_storage.stream.read(4)
    file_storage.stream.seek(0)
    return header == _PDF_MAGIC


def _sanitize_desc(value: str) -> str:
    value = str(value).strip()
    if value and value[0] in ('=', '+', '-', '@', '\t', '\r'):
        value = "'" + value
    return value


def _get_current_user():
    try:
        access_token = session.get('access_token')
        if access_token:
            user = supabase.auth.get_user(access_token)
            if user and user.user:
                return True, user.user.email, user.user.id
    except Exception:
        pass
    return False, None, None


def _get_active_chat_session(user_id: str):
    user_email_check = session.get('user_email', '')
    if user_email_check and user_email_check.lower() in OWNER_EMAILS:
        return {
            'id': f'owner_{user_id}',
            'user_id': user_id,
            'messages_used': 0,
            'statements_used': 0,
            'is_active': True,
            'input_tokens_used': 0,
            'output_tokens_used': 0,
            'input_tokens_limit': 99999999,
            'output_tokens_limit': 99999999,
        }
    try:
        result = supabase.table('chat_sessions').select('*').eq(
            'user_id', user_id
        ).eq('is_active', True).order(
            'created_at', desc=True
        ).limit(1).execute()

        if result.data:
            from datetime import datetime, timezone
            session_data = result.data[0]
            expires_at = session_data.get('expires_at')
            if expires_at:
                exp = datetime.fromisoformat(expires_at.replace('Z', '+00:00'))
                if exp < datetime.now(timezone.utc):
                    return None
            return session_data
    except Exception as e:
        logger.exception("Error checking chat session: %s", e)
    return None


def _get_locked_hash(user_id: str):
    if not user_id:
        return None
    user_email_check = session.get('user_email', '')
    if user_email_check and user_email_check.lower() in OWNER_EMAILS:
        return None
    try:
        u = supabase.table('users').select('statement_hash').eq('id', user_id).limit(1).execute()
        if u.data:
            return u.data[0].get('statement_hash')
    except Exception:
        pass
    return None


def _set_locked_hash(user_id: str, fhash: str):
    if not user_id or not fhash:
        return
    try:
        supabase.table('users').update({
            'statement_hash': fhash
        }).eq('id', user_id).execute()
    except Exception as e:
        logger.exception("Failed to set statement_hash: %s", e)


def _clear_locked_hash(user_id: str):
    if not user_id:
        return
    try:
        supabase.table('users').update({
            'statement_hash': None
        }).eq('id', user_id).execute()
    except Exception as e:
        logger.exception("Failed to clear statement_hash: %s", e)


# ═══════════════════════════════════════════════════════════
#  AUTH ROUTES
# ═══════════════════════════════════════════════════════════

@app.route('/login')
def login():
    redirect_url = ('https://www.aarogyamfin.com/auth/callback'
                    if is_prod else request.host_url + 'auth/callback')
    response = supabase.auth.sign_in_with_oauth({
        "provider": "google",
        "options":  {"redirect_to": redirect_url}
    })
    return redirect(response.url)


@app.route('/auth/callback')
def auth_callback():
    code = request.args.get('code')
    if code:
        try:
            result = supabase.auth.exchange_code_for_session({"auth_code": code})
            if result and result.session:
                session['access_token'] = result.session.access_token
                session['user_email']   = result.user.email
                session['user_id']      = result.user.id

                supabase.table('users').upsert({
                    'id':    result.user.id,
                    'email': result.user.email,
                }).execute()

                ph_track(result.user.id, event='user_signed_in', props={
                    'provider': 'google',
                    'email':    result.user.email,
                })
                logger.info("Login success: %s", result.user.email)
        except Exception as e:
            logger.exception("Auth callback error: %s", e)
    return redirect('/')


@app.route('/debug-auth')
def debug_auth():
    return {'access_token_in_session': session.get('access_token') is not None,
            'session_keys': list(session.keys())}


@app.route('/logout')
def logout():
    supabase.auth.sign_out()
    session.clear()
    return redirect('/')


# ═══════════════════════════════════════════════════════════
#  PAYMENT ROUTES
# ═══════════════════════════════════════════════════════════

@app.route('/pay')
def pay_page():
    is_logged_in, user_email, user_id = _get_current_user()
    if not is_logged_in:
        return redirect('/login')
    return render_template('payment.html',
        razorpay_key_id = RAZORPAY_KEY_ID,
        user_email      = user_email,
        is_logged_in    = is_logged_in,
    )


@app.route('/payment/create-order', methods=['POST'])
def create_order():
    is_logged_in, user_email, user_id = _get_current_user()
    if not is_logged_in:
        return jsonify({'error': 'Not logged in'}), 401

    try:
        order = rzp_client.order.create({
            'amount':          4900,
            'currency':        'INR',
            'payment_capture': 1,
        })

        supabase.table('payments').insert({
            'user_id':           user_id,
            'amount':            49,
            'status':            'pending',
            'razorpay_order_id': order['id'],
        }).execute()

        ph_track(user_id, event='payment_order_created', props={
            'order_id':   order['id'],
            'amount_inr': 49,
        })
        logger.info("Order created: %s for %s", order['id'], user_email)
        return jsonify({
            'order_id': order['id'],
            'amount':   1000,
            'currency': 'INR',
            'key_id':   RAZORPAY_KEY_ID,
            'email':    user_email,
        })

    except Exception as e:
        logger.exception("Order creation failed: %s", e)
        return jsonify({'error': 'Order creation failed'}), 500


@app.route('/payment/verify', methods=['POST'])
def verify_payment():
    is_logged_in, user_email, user_id = _get_current_user()
    if not is_logged_in:
        return jsonify({'error': 'Not logged in'}), 401

    data                = request.get_json()
    razorpay_order_id   = data.get('razorpay_order_id')
    razorpay_payment_id = data.get('razorpay_payment_id')
    razorpay_signature  = data.get('razorpay_signature')

    try:
        msg      = f"{razorpay_order_id}|{razorpay_payment_id}"
        expected = hmac.new(
            RAZORPAY_KEY_SECRET.encode(),
            msg.encode(),
            hashlib.sha256
        ).hexdigest()

        if not hmac.compare_digest(expected, razorpay_signature):
            logger.warning("Signature mismatch for order %s", razorpay_order_id)
            ph_track(user_id, event='payment_verification_failed', props={
                'order_id': razorpay_order_id,
            })
            return jsonify({'error': 'Payment verification failed'}), 400

    except Exception as e:
        logger.exception("Signature verification error: %s", e)
        return jsonify({'error': 'Verification error'}), 500

    try:
        result = supabase.table('payments').update({
            'status':              'paid',
            'razorpay_payment_id': razorpay_payment_id,
        }).eq('razorpay_order_id', razorpay_order_id).execute()

        payment_id = result.data[0]['id'] if result.data else None

        from datetime import datetime, timezone, timedelta
        expires_at = datetime.now(timezone.utc) + timedelta(hours=CHAT_SESSION_HOURS)
        supabase.table('chat_sessions').insert({
            'user_id':            user_id,
            'payment_id':         payment_id,
            'messages_used':      0,
            'is_active':          True,
            'expires_at':         expires_at.isoformat(),
            'input_tokens_limit': 75000,
            'output_tokens_limit': 10000,
            'input_tokens_used':  0,
            'output_tokens_used': 0,
        }).execute()

        _clear_locked_hash(user_id)
        session.pop('file_hash', None)
        session.pop('file_name', None)

        ph_track(user_id, event='payment_completed', props={
            'amount_inr': 49,
            'payment_id': razorpay_payment_id,
            'order_id':   razorpay_order_id,
        })
        logger.info("Payment verified + session created for %s", user_email)
        return jsonify({'success': True, 'redirect': '/chat'})

    except Exception as e:
        logger.exception("Post-payment DB error: %s", e)
        return jsonify({'error': 'Session creation failed'}), 500


# ═══════════════════════════════════════════════════════════
#  CHAT ROUTES
# ═══════════════════════════════════════════════════════════

@app.route('/chat', methods=['GET', 'POST'])
def chat_page():
    is_logged_in, user_email, user_id = _get_current_user()
    if not is_logged_in:
        return redirect('/login')

    active_session = _get_active_chat_session(user_id)
    if not active_session:
        return redirect('/pay')

    messages_left = CHAT_MESSAGE_LIMIT - active_session['messages_used']
    session_id    = str(active_session['id'])

    db_ready      = get_db(session_id) is not None
    upload_error  = None
    verify_report = None

    if request.method == 'POST' and 'pdf_file' in request.files:
        file = request.files.get('pdf_file')
        if file and file.filename:
            safe_name = secure_filename(file.filename)
            if not safe_name.lower().endswith('.pdf'):
                upload_error = 'Only PDF files accepted.'
            elif not _is_valid_pdf(file):
                upload_error = 'Invalid PDF file.'
            else:
                fhash       = _file_hash(file)
                locked_hash = _get_locked_hash(user_id)

                if locked_hash and locked_hash != fhash:
                    upload_error = NEW_STATEMENT_ERROR
                else:
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
                    file.save(filepath)
                    file_size_kb = round(os.path.getsize(filepath) / 1024, 1)
                    t0 = time.time()
                    try:
                        transactions = parse_transactions(filepath)
                        parse_time   = round(time.time() - t0, 1)
                        if transactions:
                            vr = validate_and_fix(transactions)
                            transactions = vr['transactions']
                            logger.info("Validation: %s", vr['badge'])
                            if active_session.get('statements_used', 0) >= 1 and not locked_hash:
                                upload_error = 'Statement limit reached. Please purchase a new session.'
                            else:
                                row_count     = build_index(session_id, transactions)
                                verify_report = verify_data(session_id)
                                db_ready      = True
                                _cache_set(fhash, transactions, safe_name)
                                session['file_hash'] = fhash
                                session['file_name'] = safe_name
                                if not locked_hash:
                                    _set_locked_hash(user_id, fhash)
                                supabase.table('chat_sessions').update({
                                    'statements_used': 1
                                }).eq('id', active_session['id']).execute()

                                # ── TRACK: PDF parse success ──
                                ph_track(user_id, event='pdf_parse_success', props={
                                    'context':            'chat',
                                    'file_name':          safe_name,
                                    'file_size_kb':       file_size_kb,
                                    'transaction_count':  row_count,
                                    'parse_time_seconds': parse_time,
                                    'validation_badge':   vr.get('badge', ''),
                                })
                                logger.info("Chat DB built: %d rows for session %s", row_count, session_id)
                        else:
                            upload_error = 'No transactions found in PDF.'
                            ph_track(user_id, event='pdf_parse_failed', props={
                                'context':    'chat',
                                'file_name':  safe_name,
                                'file_size_kb': file_size_kb,
                                'error_type': 'no_transactions_found',
                            })
                    except Exception as e:
                        parse_time   = round(time.time() - t0, 1)
                        upload_error = f'Parse error: {e}'
                        ph_track(user_id, event='pdf_parse_failed', props={
                            'context':       'chat',
                            'file_name':     safe_name,
                            'file_size_kb':  file_size_kb,
                            'error_type':    type(e).__name__,
                            'error_message': str(e)[:200],
                            'parse_time_seconds': parse_time,
                        })
                    finally:
                        if os.path.exists(filepath):
                            os.remove(filepath)

    elif not db_ready:
        cached_hash = session.get('file_hash')
        if cached_hash:
            cached = _cache_get(cached_hash)
            if cached and cached.get('transactions'):
                row_count     = build_index(session_id, cached['transactions'])
                verify_report = verify_data(session_id)
                db_ready      = True
                logger.info("Chat DB auto-loaded from cache: %d rows", row_count)

    return render_template('chat.html',
        is_logged_in  = is_logged_in,
        user_email    = user_email,
        messages_left = messages_left,
        session_id    = session_id,
        db_ready      = db_ready,
        verify_report = verify_report,
        upload_error  = upload_error,
    )


@app.route('/chat/message', methods=['POST'])
def chat_message():
    is_logged_in, user_email, user_id = _get_current_user()
    if not is_logged_in:
        return jsonify({'error': 'Not logged in'}), 401

    active_session = _get_active_chat_session(user_id)
    if not active_session:
        return jsonify({'error': 'No active session. Please pay to continue.'}), 403

    messages_left = CHAT_MESSAGE_LIMIT - active_session['messages_used']
    if messages_left <= 0:
        return jsonify({'error': 'Message limit reached. Please purchase a new session.'}), 403

    data       = request.get_json()
    user_msg   = (data.get('message') or '').strip()[:500]
    history    = data.get('history') or []
    session_id = str(active_session['id'])

    if not user_msg:
        return jsonify({'error': 'Empty message'}), 400

    if not get_db(session_id):
        return jsonify({'error': 'No statement loaded. Please upload a PDF first.'}), 400

    result = ai_chat(session_id, user_msg, history)

    try:
        supabase.table('chat_sessions').update({
            'messages_used':      active_session['messages_used'] + 1,
            'input_tokens_used':  active_session.get('input_tokens_used', 0) + result.get('tokens_used', 0),
            'output_tokens_used': active_session.get('output_tokens_used', 0) + result.get('output_tokens', 0),
        }).eq('id', active_session['id']).execute()
    except Exception as e:
        logger.exception("Failed to update message count: %s", e)

    ph_track(user_id, event='chat_message_sent', props={
        'messages_used':  active_session['messages_used'] + 1,
        'input_tokens':   result.get('tokens_used', 0),
        'output_tokens':  result.get('output_tokens', 0),
        'query_length':   len(user_msg),
        'reply_length':   len(result.get('reply', '')),
    })

    return jsonify({
        'reply':         result['reply'],
        'messages_left': messages_left - 1,
        'tokens_used':   result.get('tokens_used', 0),
    })


# ═══════════════════════════════════════════════════════════
#  MAIN HOME ROUTE
# ═══════════════════════════════════════════════════════════

@app.route('/', methods=['GET', 'POST'])
def home():
    ip = request.remote_addr
    if request.method == 'POST' and is_rate_limited(ip):
        abort(429)

    is_logged_in, user_email, user_id = _get_current_user()

    transaction_data  = []
    keyword           = ''
    total             = 0.0
    total_debit       = 0.0
    total_credit      = 0.0
    amounts_count     = 0
    count             = 0
    selected_filename = ''
    error_message     = ''
    parse_time        = None
    page_num          = 1
    total_pages       = 1
    total_count       = 0

    cached_hash    = session.get('file_hash')
    cached_name    = session.get('file_name', '')
    dashboard_data = None

    if request.method == 'POST':
        keyword          = request.form.get('keyword', '')[:100]
        file             = request.files.get('pdf_file')
        file_uploaded    = file and file.filename
        all_transactions = []

        if file_uploaded:
            safe_name = secure_filename(file.filename)
            if not safe_name.lower().endswith('.pdf'):
                error_message = 'Only PDF files are accepted.'
            elif len(safe_name) >= 200:
                error_message = 'Filename too long.'
            elif not _is_valid_pdf(file):
                error_message = 'Uploaded file does not appear to be a valid PDF.'
            else:
                fhash = _file_hash(file)

                locked_hash = None
                has_active  = False
                if is_logged_in:
                    active_sess = _get_active_chat_session(user_id)
                    has_active  = bool(active_sess)
                    if has_active:
                        locked_hash = _get_locked_hash(user_id)
                        if locked_hash and locked_hash != fhash:
                            error_message = NEW_STATEMENT_ERROR

                if not error_message:
                    cached = _cache_get(fhash)
                    if cached:
                        all_transactions  = cached['transactions']
                        selected_filename = cached['filename']
                        session['file_hash'] = fhash
                        session['file_name'] = selected_filename
                        if has_active and not locked_hash:
                            _set_locked_hash(user_id, fhash)
                        logger.info("Cache hit for %s", safe_name)
                    else:
                        selected_filename = safe_name
                        filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
                        file.save(filepath)
                        file_size_kb = round(os.path.getsize(filepath) / 1024, 1)
                        distinct = user_id or ip

                        # ── TRACK: Parse started ──
                        ph_track(distinct, event='pdf_parse_started', props={
                            'context':      'home',
                            'file_name':    safe_name,
                            'file_size_kb': file_size_kb,
                        })

                        try:
                            t0               = time.time()
                            all_transactions = parse_transactions(filepath)
                            parse_time       = round(time.time() - t0, 1)

                            if all_transactions:
                                vr = validate_and_fix(all_transactions)
                                all_transactions = vr['transactions']
                                session['validation_badge'] = vr['badge']
                                session['validation_color'] = vr['badge_color']
                                logger.info("Validation: %s", vr['badge'])

                            _cache_set(fhash, all_transactions, safe_name)
                            session['file_hash'] = fhash
                            session['file_name'] = safe_name
                            if has_active and not locked_hash:
                                _set_locked_hash(user_id, fhash)

                            if all_transactions:
                                # ── TRACK: Parse success ──
                                ph_track(distinct, event='pdf_parse_success', props={
                                    'context':            'home',
                                    'file_name':          safe_name,
                                    'file_size_kb':       file_size_kb,
                                    'transaction_count':  len(all_transactions),
                                    'parse_time_seconds': parse_time,
                                    'validation_badge':   vr.get('badge', ''),
                                    'is_logged_in':       is_logged_in,
                                })
                            else:
                                ph_track(distinct, event='pdf_parse_failed', props={
                                    'context':      'home',
                                    'file_name':    safe_name,
                                    'file_size_kb': file_size_kb,
                                    'error_type':   'no_transactions_found',
                                })

                            logger.info("Parsed %s → %d txns in %.1fs",
                                        safe_name, len(all_transactions), parse_time)
                        except Exception as e:
                            parse_time = round(time.time() - t0, 1)
                            logger.exception("Error parsing %s", safe_name)
                            error_message = f'Could not parse the file: {e}'
                            # ── TRACK: Parse failed ──
                            ph_track(distinct, event='pdf_parse_failed', props={
                                'context':            'home',
                                'file_name':          safe_name,
                                'file_size_kb':       file_size_kb,
                                'error_type':         type(e).__name__,
                                'error_message':      str(e)[:200],
                                'parse_time_seconds': parse_time,
                            })
                        finally:
                            if os.path.exists(filepath):
                                os.remove(filepath)

        elif cached_hash:
            cached = _cache_get(cached_hash)
            if cached:
                all_transactions  = cached['transactions']
                selected_filename = cached['filename']
            else:
                error_message = 'Session expired. Please re-upload the PDF.'
                session.pop('file_hash', None)
                session.pop('file_name', None)

        if all_transactions and not error_message:
            if keyword:
                kw_lower = keyword.lower()
                kw_clean = keyword.replace(',', '')
                filtered = [
                    t for t in all_transactions
                    if kw_lower in (t.get('desc') or '').lower()
                    or kw_lower in (t.get('date') or '').lower()
                    or kw_lower in (t.get('category') or '').lower()
                    or kw_clean in str(t.get('amount') or '')
                    or kw_clean in str(t.get('balance') or '')
                ]
            else:
                filtered = all_transactions

            for t in filtered:
                if t.get('amount'):
                    amounts_count += 1
                    if t.get('type') == 'DR':
                        total_debit  += t['amount']
                    else:
                        total_credit += t['amount']

            total       = total_debit + total_credit
            total_count = len(filtered)
            page_num    = int(request.form.get('page', 1))
            total_pages = max(1, (total_count + PAGE_SIZE - 1) // PAGE_SIZE)
            page_num    = max(1, min(page_num, total_pages))
            start            = (page_num - 1) * PAGE_SIZE
            transaction_data = filtered[start: start + PAGE_SIZE]
            count            = total_count

            try:
                dashboard_data = run_dashboard(all_transactions[:300])
            except Exception as e:
                logger.exception("Dashboard POST error: %s", e)
                dashboard_data = None

    elif request.method == 'GET' and cached_hash:
        selected_filename = cached_name
        cached = _cache_get(cached_hash)
        if cached and cached.get('transactions'):
            try:
                dashboard_data = run_dashboard(cached['transactions'][:300])
            except Exception as e:
                logger.exception("Dashboard GET error: %s", e)
                dashboard_data = None

    return render_template(
        'index.html',
        transaction_data  = transaction_data,
        keyword           = keyword,
        total             = total,
        total_debit       = total_debit,
        total_credit      = total_credit,
        count             = count,
        amounts_count     = amounts_count,
        selected_filename = selected_filename,
        error_message     = error_message,
        parse_time        = parse_time,
        has_cached_file   = bool(session.get('file_hash')),
        cached_filename   = session.get('file_name', ''),
        is_logged_in      = is_logged_in,
        user_email        = user_email,
        page_num          = page_num,
        total_pages       = total_pages,
        total_count       = total_count,
        dashboard_data    = dashboard_data,
    )


@app.route('/clear', methods=['POST'])
def clear_cache():
    fhash = session.pop('file_hash', None)
    session.pop('file_name', None)
    if fhash:
        with _CACHE_LOCK:
            _CACHE.pop(fhash, None)
    return ('', 204)


@app.route('/sitemap.xml')
def sitemap():
    pages = [
        ('https://aarogyamfin.com/',        '2026-04-08'),
        ('https://aarogyamfin.com/about',   '2026-04-08'),
        ('https://aarogyamfin.com/accuracy','2026-04-07'),
        ('https://aarogyamfin.com/privacy', '2026-04-08'),
        ('https://aarogyamfin.com/terms',   '2026-04-08'),
        ('https://aarogyamfin.com/contact', '2026-04-08'),
        ('https://aarogyamfin.com/blog', '2026-06-19'),
        ('https://aarogyamfin.com/bank-statement-analyzer', '2026-06-23'),
        ('https://aarogyamfin.com/bank-statement-analyzer-for-ca', '2026-06-23'),
        ('https://aarogyamfin.com/bank-statement-analyzer-for-dsa', '2026-06-23'),
        ('https://aarogyamfin.com/bank-statement-analyzer-for-nbfc', '2026-06-23'),
        ('https://aarogyamfin.com/bank-statement-to-excel', '2026-06-23'),
        ('https://aarogyamfin.com/ai-bank-statement-analyzer', '2026-06-23'),
        ('https://aarogyamfin.com/pdf-to-excel-bank-statement', '2026-06-23'),
        ('https://aarogyamfin.com/bank-statement-analysis', '2026-06-23'),
        ('https://aarogyamfin.com/bank-statement-analysis-software', '2026-06-23'),
        ('https://aarogyamfin.com/bank-statement-ocr', '2026-06-23'),
        ('https://aarogyamfin.com/bank-statement-converter', '2026-06-23'),
        ('https://aarogyamfin.com/bank-pdf-to-excel', '2026-06-23'),
        ('https://aarogyamfin.com/bank-statement-excel-converter', '2026-06-23'),
        ('https://aarogyamfin.com/financial-statement-analyzer', '2026-06-23'),
        ('https://aarogyamfin.com/statement-parser', '2026-06-23'),
        ('https://aarogyamfin.com/statement-analysis-software', '2026-06-23'),
        ('https://aarogyamfin.com/bank-data-extraction', '2026-06-23'),
        ('https://aarogyamfin.com/automatic-bank-statement-analysis', '2026-06-23'),
        ('https://aarogyamfin.com/ai-bank-statement-parser', '2026-06-23'),
        ('https://aarogyamfin.com/bank-statement-analysis-tool', '2026-06-23'),
        ('https://aarogyamfin.com/pdf-bank-parser', '2026-06-23'),
        ('https://aarogyamfin.com/bank-statement-parser', '2026-06-23'),
        ('https://aarogyamfin.com/online-bank-statement-analyzer', '2026-06-23'),
        ('https://aarogyamfin.com/sbi-bank-statement-analyzer', '2026-06-23'),
        ('https://aarogyamfin.com/hdfc-bank-statement-analyzer', '2026-06-23'),
        ('https://aarogyamfin.com/icici-bank-statement-analyzer', '2026-06-23'),
        ('https://aarogyamfin.com/axis-bank-statement-analyzer', '2026-06-23'),
        ('https://aarogyamfin.com/kotak-bank-statement-analyzer', '2026-06-23'),
        ('https://aarogyamfin.com/blog/how-to-analyze-bank-statement-for-itr', '2026-06-19'),
    ]
    xml = ['<?xml version="1.0" encoding="UTF-8"?>',
           '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
    for url, lastmod in pages:
        xml.append(f'  <url><loc>{url}</loc><lastmod>{lastmod}</lastmod></url>')
    xml.append('</urlset>')
    return Response('\n'.join(xml), mimetype='application/xml')


@app.route('/llms.txt')
def llms():
    content = """# AarogyamFin — AI Bank Statement Analyzer
# https://aarogyamfin.com

## Product
AarogyamFin is an AI-powered Indian bank statement analyzer.
Users upload bank statement PDFs and instantly get transaction analysis, keyword search, Excel export, and AI chat insights.

## Supported Banks
SBI, HDFC, Kotak, Canara, Axis, Punjab National Bank, Bank of Baroda, ICICI, and all major Indian banks.

## Features
- PDF to Excel conversion
- Transaction search by keyword
- Credit/Debit categorization
- AI chat for financial insights
- Dashboard with ITR, loan eligibility, compliance analysis
- Mobile app (Android)

## Pricing
Free basic analysis. AI Chat sessions from ₹10.

## Contact
aarogyamfin@gmail.com

## URLs
Homepage: https://aarogyamfin.com
About: https://aarogyamfin.com/about
Privacy: https://aarogyamfin.com/privacy
Terms: https://aarogyamfin.com/terms
"""
    return Response(content, mimetype='text/plain')

@app.route('/robots.txt')
def robots():
    content = """User-agent: *
Allow: /

User-agent: Googlebot
Allow: /

User-agent: Google-Extended
Allow: /

User-agent: GPTBot
Allow: /

User-agent: ClaudeBot
Allow: /

User-agent: Applebot-Extended
Allow: /

User-agent: meta-externalagent
Allow: /

User-agent: CCBot
Disallow: /

User-agent: Bytespider
Disallow: /

Sitemap: https://aarogyamfin.com/sitemap.xml
"""
    return Response(content, mimetype='text/plain')

@app.route('/about')
def about():
    return render_template('about.html')


# ═══════════════════════════════════════════════════════════
#  SEO LANDING PAGES
# ═══════════════════════════════════════════════════════════

@app.route('/bank-statement-analyzer')
def bank_statement_analyzer():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/bank-statement-analyzer-for-ca')
def bsa_for_ca():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/sbi-bank-statement-analyzer')
def sbi_bsa():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/hdfc-bank-statement-analyzer')
def hdfc_bsa():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/icici-bank-statement-analyzer')
def icici_bsa():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/axis-bank-statement-analyzer')
def axis_bsa():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/kotak-bank-statement-analyzer')
def kotak_bsa():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/pnb-bank-statement-analyzer')
def pnb_bsa():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/bob-bank-statement-analyzer')
def bob_bsa():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/canara-bank-statement-analyzer')
def canara_bsa():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/union-bank-statement-analyzer')
def union_bsa():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/bank-statement-analyzer-for-dsa')
def bsa_for_dsa():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/bank-statement-analyzer-for-nbfc')
def bsa_for_nbfc():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/privacy')
def privacy():
    return render_template('privacy.html')

@app.route('/terms')
def terms():
    return render_template('terms.html')

@app.route('/bank-statement-to-excel')
def bank_statement_to_excel():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_to_excel.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/ai-bank-statement-analyzer')
def ai_bank_statement_analyzer():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('ai_bank_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/bank-statement-parser')
def bank_statement_parser():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_parser.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/online-bank-statement-analyzer')
def online_bank_statement_analyzer():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('online_bank_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/pdf-to-excel-bank-statement')
def pdf_to_excel_bank_statement():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('pdf_to_excel_bank_statement.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/bank-statement-analysis')
def bank_statement_analysis():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analysis.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/bank-statement-analysis-software')
def bank_statement_analysis_software():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analysis_software.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/bank-statement-ocr')
def bank_statement_ocr():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_ocr.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/bank-statement-converter')
def bank_statement_converter():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_converter.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/bank-pdf-to-excel')
def bank_pdf_to_excel():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_pdf_to_excel.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/bank-statement-excel-converter')
def bank_statement_excel_converter():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_excel_converter.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/financial-statement-analyzer')
def financial_statement_analyzer():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('financial_statement_analyzer.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/statement-parser')
def statement_parser():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('statement_parser.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/statement-analysis-software')
def statement_analysis_software():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('statement_analysis_software.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/bank-data-extraction')
def bank_data_extraction():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_data_extraction.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/automatic-bank-statement-analysis')
def automatic_bank_statement_analysis():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('automatic_bank_statement_analysis.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/ai-bank-statement-parser')
def ai_bank_statement_parser():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('ai_bank_statement_parser.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/bank-statement-analysis-tool')
def bank_statement_analysis_tool():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('bank_statement_analysis_tool.html', is_logged_in=is_logged_in, user_email=user_email)

@app.route('/pdf-bank-parser')
def pdf_bank_parser():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('pdf_bank_parser.html', is_logged_in=is_logged_in, user_email=user_email)


# ═══════════════════════════════════════════════════════════
#  ACCURACY ROUTES
# ═══════════════════════════════════════════════════════════

@app.route('/accuracy', methods=['GET'])
def accuracy_page():
    return render_template('accuracy.html', error=None)


@app.route('/accuracy', methods=['POST'])
def accuracy_check():
    ip = request.remote_addr
    if is_rate_limited(ip):
        abort(429)

    file = request.files.get('pdf_file')
    if not file or not file.filename:
        return render_template('accuracy.html', error="No file uploaded.")

    safe_name = secure_filename(file.filename)
    if not safe_name.lower().endswith('.pdf'):
        return render_template('accuracy.html', error="Only PDF files accepted.")
    if len(safe_name) >= 200:
        return render_template('accuracy.html', error="Filename too long.")
    if not _is_valid_pdf(file):
        return render_template('accuracy.html', error="Invalid PDF file.")

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
    file.save(filepath)

    try:
        t0           = time.time()
        transactions = parse_transactions(filepath)
        parse_time   = round(time.time() - t0, 1)
        logger.info("Accuracy test: parsed %s → %d txns in %.1fs",
                    safe_name, len(transactions), parse_time)
    except Exception as e:
        logger.exception("Accuracy test parse error: %s", safe_name)
        return render_template('accuracy.html', error=f"Parse error: {e}")
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)

    if not transactions:
        return render_template('accuracy.html', error="No transactions found in this PDF.")

    report = run_accuracy_check(transactions)
    return render_template('accuracy.html', filename=safe_name, parse_time=parse_time, **report)


# ═══════════════════════════════════════════════════════════
#  EXPORT ROUTE (Excel)
# ═══════════════════════════════════════════════════════════

@app.route('/export', methods=['POST'])
def export():
    ip = request.remote_addr
    if is_rate_limited(ip):
        abort(429)

    is_logged_in, user_email, user_id = _get_current_user()

    dates    = request.form.getlist('dates')
    descs    = request.form.getlist('descs')
    amounts  = request.form.getlist('amounts')
    balances = request.form.getlist('balances')
    types    = request.form.getlist('types')
    cats     = request.form.getlist('categories')
    keyword  = request.form.get('keyword', '')[:100]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(["#", "Date", "Description", "Type", "Amount (₹)", "Balance (₹)", "Category"])

    for i, (d, desc, a, b, tp, cat) in enumerate(
            zip(dates, descs, amounts, balances, types,
                cats or [''] * len(dates)), 1):
        try:    amt = float(a) if a else ''
        except: amt = ''
        try:    bal = float(b) if b else ''
        except: bal = ''
        ws.append([i, d, _sanitize_desc(desc), tp, amt, bal, cat])

    # ── TRACK: Excel export ──
    ph_track(user_id or ip, event='excel_exported', props={
        'context':           'home',
        'transaction_count': len(dates),
        'keyword_used':      bool(keyword),
    })

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_kw = secure_filename(keyword) if keyword else 'transactions'
    return send_file(output, download_name=f'transactions_{safe_kw}.xlsx', as_attachment=True)


# ═══════════════════════════════════════════════════════════
#  DASHBOARD ROUTES
# ═══════════════════════════════════════════════════════════

@app.route('/dashboard', methods=['GET'])
def dashboard_page():
    is_logged_in, user_email, user_id = _get_current_user()
    if not is_logged_in:
        return redirect('/login')
    active = _get_active_chat_session(user_id)
    if not active:
        return redirect('/pay')

    data        = None
    cached_hash = session.get('file_hash')
    locked_hash = _get_locked_hash(user_id)

    if cached_hash and (not locked_hash or cached_hash == locked_hash):
        cached = _cache_get(cached_hash)
        if cached:
            data = run_dashboard(cached['transactions'])

    return render_template(
        'dashboard.html',
        data         = data,
        is_logged_in = is_logged_in,
        user_email   = user_email,
    )


@app.route('/dashboard', methods=['POST'])
def dashboard_analyze():
    ip = request.remote_addr
    if is_rate_limited(ip):
        abort(429)

    is_logged_in, user_email, user_id = _get_current_user()
    if not is_logged_in:
        return redirect('/login')
    active = _get_active_chat_session(user_id)
    if not active:
        return redirect('/pay')

    data          = None
    error_message = ''

    if request.form.get('use_cached'):
        cached_hash = session.get('file_hash')
        locked_hash = _get_locked_hash(user_id)
        if cached_hash and (not locked_hash or cached_hash == locked_hash):
            cached = _cache_get(cached_hash)
            if cached:
                data = run_dashboard(cached['transactions'])
            else:
                error_message = 'Session expired. Please re-upload.'
        else:
            error_message = NEW_STATEMENT_ERROR
        return render_template('dashboard.html',
            data          = data,
            error_message = error_message,
            is_logged_in  = is_logged_in,
            user_email    = user_email,
        )

    file = request.files.get('pdf_file')
    logger.info("DEBUG: session file_hash=%s, user_id=%s", session.get('file_hash'), user_id)
    if not file or not file.filename:
        return render_template('dashboard.html',
            data          = None,
            error_message = 'No file uploaded.',
            is_logged_in  = is_logged_in,
            user_email    = user_email,
        )

    safe_name = secure_filename(file.filename)
    if not safe_name.lower().endswith('.pdf'):
        error_message = 'Only PDF files accepted.'
    elif len(safe_name) >= 200:
        error_message = 'Filename too long.'
    elif not _is_valid_pdf(file):
        error_message = 'Invalid PDF file.'
    else:
        fhash       = _file_hash(file)
        locked_hash = _get_locked_hash(user_id)

        if locked_hash and locked_hash != fhash:
            return render_template('dashboard.html',
                data          = None,
                error_message = NEW_STATEMENT_ERROR,
                is_logged_in  = is_logged_in,
                user_email    = user_email,
            )

        cached = _cache_get(fhash)

        if cached:
            transactions = cached['transactions']
            session['file_hash'] = fhash
            session['file_name'] = cached.get('filename', safe_name)
            if not locked_hash:
                _set_locked_hash(user_id, fhash)
            logger.info("Dashboard cache hit for %s", safe_name)
        else:
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
            file.save(filepath)
            file_size_kb = round(os.path.getsize(filepath) / 1024, 1)
            t0 = time.time()
            try:
                transactions = parse_transactions(filepath)
                parse_time   = round(time.time() - t0, 1)
                _cache_set(fhash, transactions, safe_name)
                session['file_hash'] = fhash
                session['file_name'] = safe_name
                if not locked_hash:
                    _set_locked_hash(user_id, fhash)

                ph_track(user_id, event='pdf_parse_success', props={
                    'context':            'dashboard',
                    'file_name':          safe_name,
                    'file_size_kb':       file_size_kb,
                    'transaction_count':  len(transactions),
                    'parse_time_seconds': parse_time,
                })
                logger.info("Dashboard parsed %s → %d txns in %.1fs",
                            safe_name, len(transactions), parse_time)
            except Exception as e:
                parse_time    = round(time.time() - t0, 1)
                logger.exception("Dashboard parse error: %s", safe_name)
                error_message = f'Could not parse: {e}'
                transactions  = []
                ph_track(user_id, event='pdf_parse_failed', props={
                    'context':            'dashboard',
                    'file_name':          safe_name,
                    'file_size_kb':       file_size_kb,
                    'error_type':         type(e).__name__,
                    'error_message':      str(e)[:200],
                    'parse_time_seconds': parse_time,
                })
            finally:
                if os.path.exists(filepath):
                    os.remove(filepath)

        if transactions and not error_message:
            data = run_dashboard(transactions)

    return render_template(
        'dashboard.html',
        data          = data,
        error_message = error_message,
        is_logged_in  = is_logged_in,
        user_email    = user_email,
    )


# ═══════════════════════════════════════════════════════════
#  DASHBOARD EXCEL EXPORT
# ═══════════════════════════════════════════════════════════

@app.route('/dashboard/export', methods=['GET'])
def dashboard_export():
    ip = request.remote_addr
    if is_rate_limited(ip):
        abort(429)

    is_logged_in, user_email, user_id = _get_current_user()
    if not is_logged_in:
        return redirect('/login')

    cached_hash = session.get('file_hash')
    if not cached_hash:
        return "No data available. Please upload a PDF first.", 400

    locked_hash = _get_locked_hash(user_id)
    if locked_hash and cached_hash != locked_hash:
        return NEW_STATEMENT_ERROR, 403

    cached = _cache_get(cached_hash)
    if not cached:
        return "Session expired. Please re-upload.", 400

    data     = run_dashboard(cached['transactions'])
    filename = session.get('file_name', 'statement')

    # ── TRACK: Excel dashboard export ──
    ph_track(user_id, event='dashboard_excel_exported', props={
        'file_name':         filename,
        'transaction_count': len(cached['transactions']),
    })

    d_exp  = data['expense']
    d_itr  = data['itr']
    d_aud  = data['audit']
    d_loan = data['loan']
    d_comp = data['compliance']

    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    COLOR_DARK       = '0D1117'
    COLOR_GOLD       = 'C9A84C'
    COLOR_GOLD_LIGHT = 'F0D98A'
    COLOR_GREEN      = '4ADE80'
    COLOR_RED        = 'F87171'
    COLOR_YELLOW     = 'FBBF24'
    COLOR_HEADER_BG  = '1A1F2E'
    COLOR_ROW_ALT    = '0F1420'
    COLOR_WHITE      = 'E8E4D9'
    COLOR_MUTED      = '8A8070'

    def hdr_fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    def hdr_font(hex_color='E8E4D9', bold=True, size=10):
        return Font(color=hex_color, bold=bold, size=size, name='Calibri')

    def cell_font(hex_color='E8E4D9', bold=False, size=9):
        return Font(color=hex_color, bold=bold, size=size, name='Calibri')

    def thin_border():
        s = Side(style='thin', color='2A2F3E')
        return Border(left=s, right=s, top=s, bottom=s)

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def write_title(ws, title, subtitle=''):
        ws.sheet_view.showGridLines = False
        ws['A1'] = title
        ws['A1'].font = Font(color=COLOR_GOLD, bold=True, size=14, name='Calibri')
        ws['A1'].fill = hdr_fill(COLOR_DARK)
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 32
        if subtitle:
            ws['A2'] = subtitle
            ws['A2'].font = Font(color=COLOR_MUTED, bold=False, size=9, name='Calibri')
            ws['A2'].fill = hdr_fill(COLOR_DARK)
            ws.row_dimensions[2].height = 18
            return 4
        return 3

    def write_section_header(ws, row, label, col_count=6):
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(color=COLOR_GOLD_LIGHT, bold=True, size=9, name='Calibri')
        ws.cell(row=row, column=1).fill = hdr_fill('1A2030')
        ws.cell(row=row, column=1).alignment = Alignment(vertical='center')
        ws.row_dimensions[row].height = 20
        for c in range(2, col_count + 1):
            ws.cell(row=row, column=c).fill = hdr_fill('1A2030')

    def write_table_header(ws, row, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = hdr_font(COLOR_GOLD)
            cell.fill = hdr_fill(COLOR_HEADER_BG)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border()
        ws.row_dimensions[row].height = 20

    def write_data_row(ws, row, values, colors=None):
        fill_color = COLOR_ROW_ALT if row % 2 == 0 else COLOR_DARK
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            color = (colors[c-1] if colors and c <= len(colors) else COLOR_WHITE)
            cell.font = cell_font(color)
            cell.fill = hdr_fill(fill_color)
            cell.alignment = Alignment(vertical='center', wrap_text=False)
            cell.border = thin_border()
        ws.row_dimensions[row].height = 16

    def write_kv(ws, row, label, value, val_color=COLOR_WHITE):
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = cell_font(COLOR_MUTED)
        ws.cell(row=row, column=1).fill = hdr_fill(COLOR_DARK)
        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=2).font = cell_font(val_color, bold=True)
        ws.cell(row=row, column=2).fill = hdr_fill(COLOR_DARK)
        ws.row_dimensions[row].height = 15

    # Sheet 1: Overview
    ws1 = wb.active
    ws1.title = 'Overview'
    r = write_title(ws1, '  AarogyamFin — Financial Intelligence Report', f'  Generated for: {filename}')
    write_section_header(ws1, r, '  SUMMARY METRICS', 3)
    r += 1
    write_table_header(ws1, r, ['Metric', 'Value', ''])
    r += 1
    for label, value, color in [
        ('Total Transactions',   data['total_txns'],                                                 COLOR_WHITE),
        ('Total Credits',        f'Rs {data["total_cr"]:,.0f}',                                      COLOR_GREEN),
        ('Total Debits',         f'Rs {data["total_dr"]:,.0f}',                                      COLOR_RED),
        ('Real Income',          f'Rs {d_itr["real_income_total"]:,.0f}',                            COLOR_GREEN),
        ('Loan Disbursals',      f'Rs {d_itr["loan_disbursal_total"]:,.0f}',                         COLOR_RED),
        ('Family Transfers',     f'Rs {d_itr.get("family_transfer_total", 0):,.0f}',                 COLOR_YELLOW),
        ('Avg Monthly Income',   f'Rs {d_loan["avg_monthly_credit"]:,.0f}',                          COLOR_GREEN),
        ('Avg Monthly Spend',    f'Rs {d_loan["avg_monthly_debit"]:,.0f}',                           COLOR_RED),
        ('Avg Balance',          f'Rs {d_loan["avg_balance"]:,.0f}',                                 COLOR_GOLD_LIGHT),
        ('FOIR',                 f'{d_loan["foir"]}%',                                               COLOR_YELLOW),
        ('Loan Eligible',        f'Rs {d_loan["loan_eligible"]:,.0f}',                               COLOR_GREEN),
        ('Credit Profile',       d_loan['credit_indicator'],                                         COLOR_GOLD_LIGHT),
        ('Reconciliation Score', f'{d_aud["reconciliation_score"]}/100',                             COLOR_GREEN),
        ('Balance Mismatches',   d_aud['mismatch_count'],                                            COLOR_RED if d_aud['mismatch_count'] > 0 else COLOR_GREEN),
        ('EMIs Detected',        d_aud['emi_count'],                                                 COLOR_YELLOW),
        ('Compliance Risk',      d_comp['risk_level'],                                               COLOR_RED if d_comp['risk_level'] == 'High' else COLOR_YELLOW if d_comp['risk_level'] == 'Medium' else COLOR_GREEN),
        ('Suggested ITR Form',   d_itr['suggested_itr'].split('(')[0].strip(),                       COLOR_GOLD_LIGHT),
        ('80C Deductions',       f'Rs {d_itr["section_80c_total"]:,.0f}',                            COLOR_YELLOW),
        ('80D Deductions',       f'Rs {d_itr["section_80d_total"]:,.0f}',                            COLOR_YELLOW),
        ('Form 61A Required',    'YES' if d_comp['form_61a_required'] else 'NO',                     COLOR_RED if d_comp['form_61a_required'] else COLOR_GREEN),
    ]:
        write_kv(ws1, r, f'  {label}', value, color)
        r += 1
    set_col_widths(ws1, [32, 28, 10])

    # ── SHEET 2: Expenses ──
    ws2 = wb.create_sheet('Expenses')
    r = write_title(ws2, '  Module 01 — Expense Categorization', '  Business vs Personal · GST-eligible transactions')
    write_section_header(ws2, r, '  EXPENSE SUMMARY', 4)
    r += 1
    write_table_header(ws2, r, ['Category', 'Amount (Rs)', 'Count', '% of Debits'])
    r += 1
    for row_data in [
        ('Business Expenses',  d_exp['business_total'],     len(d_exp['business']),     d_exp['business_pct']),
        ('Personal Expenses',  d_exp['personal_total'],     len(d_exp['personal']),     d_exp['personal_pct']),
        ('GST Input Eligible', d_exp['gst_eligible_total'], len(d_exp['gst_eligible']), round(d_exp['gst_eligible_total'] / d_exp['total_debits'] * 100, 1) if d_exp['total_debits'] else 0),
        ('Mixed/Unclassified', d_exp['mixed_total'],        len(d_exp['mixed']),        round(d_exp['mixed_total'] / d_exp['total_debits'] * 100, 1) if d_exp['total_debits'] else 0),
    ]:
        write_data_row(ws2, r, [f'  {row_data[0]}', f'Rs {row_data[1]:,.0f}', row_data[2], f'{row_data[3]:.1f}%'])
        r += 1
    r += 1
    write_section_header(ws2, r, '  CATEGORY BREAKDOWN', 4)
    r += 1
    write_table_header(ws2, r, ['Category', 'Amount (Rs)', '% of Total', ''])
    r += 1
    for cat, amt in d_exp['category_totals'].items():
        pct = round(amt / d_exp['total_debits'] * 100, 1) if d_exp['total_debits'] else 0
        write_data_row(ws2, r, [f'  {cat}', f'Rs {amt:,.0f}', f'{pct}%', ''])
        r += 1
    r += 1
    write_section_header(ws2, r, '  GST INPUT CREDIT ELIGIBLE TRANSACTIONS', 4)
    r += 1
    write_table_header(ws2, r, ['Date', 'Description', 'Amount (Rs)', 'Type'])
    r += 1
    for t in d_exp['gst_eligible']:
        write_data_row(ws2, r, [t.get('date',''), t.get('desc','')[:80], f'Rs {t.get("amount",0):,.2f}', 'Business'])
        r += 1
    set_col_widths(ws2, [14, 55, 20, 16])

    # ── SHEET 3: ITR ──
    ws3 = wb.create_sheet('ITR Tax')
    r = write_title(ws3, '  Module 02 — ITR / Tax Filing', f'  Suggested: {d_itr["suggested_itr"].split("(")[0].strip()}')
    write_section_header(ws3, r, '  INCOME SUMMARY', 3)
    r += 1
    write_table_header(ws3, r, ['Income Source', 'Total Amount (Rs)', 'Transactions'])
    r += 1
    colors_map = [COLOR_WHITE, COLOR_WHITE, COLOR_WHITE, COLOR_WHITE, COLOR_RED, COLOR_YELLOW, COLOR_GREEN]
    for i, (src, amt, cnt) in enumerate([
        ('Salary',             d_itr['salary_total'],                  len(d_itr['income_sources'].get('salary', []))),
        ('Freelance/Business', d_itr['freelance_total'],               len(d_itr['income_sources'].get('freelance', []))),
        ('Interest Income',    d_itr['interest_total'],                len(d_itr['income_sources'].get('interest', []))),
        ('Other Credits',      d_itr.get('other_credits', 0),          0),
        ('Loan Disbursals',    d_itr['loan_disbursal_total'],          0),
        ('Family Transfers',   d_itr.get('family_transfer_total', 0), 0),
        ('REAL INCOME TOTAL',  d_itr['real_income_total'],             0),
    ]):
        clr = colors_map[i]
        write_data_row(ws3, r, [f'  {src}', f'Rs {amt:,.0f}', cnt if cnt else '—'], colors=[clr, clr, clr])
        r += 1
    r += 1
    write_section_header(ws3, r, '  TAX DEDUCTIONS', 3)
    r += 1
    write_table_header(ws3, r, ['Section', 'Total (Rs)', 'Max Limit'])
    r += 1
    write_data_row(ws3, r, ['  80C (LIC/PF/PPF)', f'Rs {d_itr["section_80c_total"]:,.0f}', 'Rs 1,50,000'])
    r += 1
    write_data_row(ws3, r, ['  80D (Health Ins)',  f'Rs {d_itr["section_80d_total"]:,.0f}', 'Rs 25,000'])
    r += 1
    r += 1
    write_section_header(ws3, r, '  HIGH VALUE CREDITS >= Rs 1 LAKH', 3)
    r += 1
    write_table_header(ws3, r, ['Date', 'Description', 'Amount (Rs)'])
    r += 1
    for t in d_itr['high_value_credits']:
        write_data_row(ws3, r, [t.get('date',''), t.get('desc','')[:80], f'Rs {t.get("amount",0):,.0f}'],
                       colors=[COLOR_MUTED, COLOR_WHITE, COLOR_GREEN])
        r += 1
    set_col_widths(ws3, [30, 55, 22])

    # ── SHEET 4: Audit ──
    ws4 = wb.create_sheet('Audit')
    r = write_title(ws4, '  Module 03 — Audit & Reconciliation', f'  Score: {d_aud["reconciliation_score"]}/100')
    write_section_header(ws4, r, '  AUDIT SUMMARY', 3)
    r += 1
    for label, value, color in [
        ('Reconciliation Score', f'{d_aud["reconciliation_score"]}/100',  COLOR_GREEN if d_aud['reconciliation_score'] >= 80 else COLOR_YELLOW),
        ('Balance Mismatches',   d_aud['mismatch_count'],                 COLOR_RED if d_aud['mismatch_count'] > 0 else COLOR_GREEN),
        ('Bounced Transactions', d_aud['bounce_count'],                   COLOR_RED if d_aud['bounce_count'] > 0 else COLOR_GREEN),
        ('EMIs Detected',        d_aud['emi_count'],                      COLOR_YELLOW),
        ('EMI Total Outflow',    f'Rs {d_aud["total_emi_outflow"]:,.0f}', COLOR_RED),
        ('EMI to Income Ratio',  f'{d_aud["emi_to_income_ratio"]}%',      COLOR_YELLOW),
        ('Duplicate Suspects',   d_aud['duplicate_count'],                COLOR_YELLOW if d_aud['duplicate_count'] > 0 else COLOR_GREEN),
    ]:
        write_kv(ws4, r, f'  {label}', value, color)
        r += 1
    r += 1
    write_section_header(ws4, r, '  EMI / LOAN OUTFLOWS', 4)
    r += 1
    write_table_header(ws4, r, ['Date', 'Description', 'Amount (Rs)', 'Type'])
    r += 1
    for t in d_aud['emis']:
        write_data_row(ws4, r, [t.get('date',''), t.get('desc','')[:80], f'Rs {t.get("amount",0):,.2f}', 'EMI/Loan'],
                       colors=[COLOR_MUTED, COLOR_WHITE, COLOR_RED, COLOR_YELLOW])
        r += 1
    if d_aud['bounced']:
        r += 1
        write_section_header(ws4, r, '  BOUNCED TRANSACTIONS', 3)
        r += 1
        write_table_header(ws4, r, ['Date', 'Description', 'Amount (Rs)'])
        r += 1
        for t in d_aud['bounced']:
            write_data_row(ws4, r, [t.get('date',''), t.get('desc','')[:80], f'Rs {t.get("amount",0):,.2f}'],
                           colors=[COLOR_MUTED, COLOR_WHITE, COLOR_RED])
            r += 1
    set_col_widths(ws4, [14, 55, 20, 16])

    # ── SHEET 5: Loan ──
    ws5 = wb.create_sheet('Loan')
    r = write_title(ws5, '  Module 04 — Loan & Credit Assessment', f'  Profile: {d_loan["credit_indicator"]}')
    write_section_header(ws5, r, '  LOAN METRICS', 3)
    r += 1
    for label, value, color in [
        ('Credit Profile',          d_loan['credit_indicator'],                                               COLOR_GREEN if d_loan['credit_color'] == 'green' else COLOR_RED if d_loan['credit_color'] == 'red' else COLOR_YELLOW),
        ('FOIR',                    f'{d_loan["foir"]}%',                                                     COLOR_RED if d_loan['foir'] > 50 else COLOR_YELLOW if d_loan['foir'] > 35 else COLOR_GREEN),
        ('DSCR',                    f'{d_loan["dscr"]}x' if d_loan.get('dscr') else 'N/A',                   COLOR_WHITE),
        ('Avg Monthly Income',      f'Rs {d_loan["avg_monthly_credit"]:,.0f}',                                COLOR_GREEN),
        ('Avg Monthly Spend',       f'Rs {d_loan["avg_monthly_debit"]:,.0f}',                                 COLOR_RED),
        ('Avg Balance',             f'Rs {d_loan["avg_balance"]:,.0f}',                                       COLOR_GOLD_LIGHT),
        ('Min Balance',             f'Rs {d_loan["min_balance"]:,.0f}',                                       COLOR_YELLOW),
        ('Negative Balance Months', d_loan['negative_months'],                                                COLOR_RED if d_loan['negative_months'] > 0 else COLOR_GREEN),
        ('Remaining EMI Capacity',  f'Rs {d_loan["remaining_emi_capacity"]:,.0f}/mo',                         COLOR_GREEN),
        ('Estimated Loan Eligible', f'Rs {d_loan["loan_eligible"]:,.0f}',                                     COLOR_GREEN if d_loan['loan_eligible'] > 0 else COLOR_RED),
        ('Months Analyzed',         d_loan['months_analyzed'],                                                COLOR_WHITE),
    ]:
        write_kv(ws5, r, f'  {label}', value, color)
        r += 1
    r += 1
    write_section_header(ws5, r, '  MONTHLY CASH FLOW', 4)
    r += 1
    write_table_header(ws5, r, ['Month', 'Credits (Rs)', 'Debits (Rs)', 'Net (Rs)'])
    r += 1
    for month in sorted(d_loan['monthly_data'].keys()):
        md  = d_loan['monthly_data'][month]
        net = md['credits'] - md['debits']
        write_data_row(ws5, r, [month, f'Rs {md["credits"]:,.0f}', f'Rs {md["debits"]:,.0f}', f'Rs {net:,.0f}'],
                       colors=[COLOR_MUTED, COLOR_GREEN, COLOR_RED, COLOR_GREEN if net >= 0 else COLOR_RED])
        r += 1
    set_col_widths(ws5, [18, 22, 22, 22])

    # ── SHEET 6: Compliance ──
    ws6 = wb.create_sheet('Compliance')
    r = write_title(ws6, '  Module 05 — Compliance Reporting', f'  Risk: {d_comp["risk_level"]} · Score: {d_comp["risk_score"]}')
    write_section_header(ws6, r, '  COMPLIANCE SUMMARY', 3)
    r += 1
    for label, value, color in [
        ('Risk Level',             d_comp['risk_level'],                                                       COLOR_RED if d_comp['risk_level'] == 'High' else COLOR_YELLOW if d_comp['risk_level'] == 'Medium' else COLOR_GREEN),
        ('Risk Score',             d_comp['risk_score'],                                                       COLOR_WHITE),
        ('High Value Txns (>=2L)', d_comp['high_value_count'],                                                 COLOR_YELLOW if d_comp['high_value_count'] > 5 else COLOR_GREEN),
        ('Cash Transactions',      d_comp['cash_count'],                                                       COLOR_WHITE),
        ('STR Candidates',         d_comp['str_count'],                                                        COLOR_RED if d_comp['str_count'] > 0 else COLOR_GREEN),
        ('Structuring Suspects',   d_comp['structured_count'],                                                 COLOR_RED if d_comp['structured_count'] > 0 else COLOR_GREEN),
        ('Daily Limit Breaches',   len(d_comp['daily_breaches']),                                              COLOR_RED if d_comp['daily_breaches'] else COLOR_GREEN),
        ('Annual Cash Total',      f'Rs {d_comp["annual_cash_total"]:,.0f}',                                   COLOR_YELLOW),
        ('Form 61A / SFT',         'REQUIRED' if d_comp['form_61a_required'] else 'Not Required',              COLOR_RED if d_comp['form_61a_required'] else COLOR_GREEN),
    ]:
        write_kv(ws6, r, f'  {label}', value, color)
        r += 1
    if d_comp['high_value_txns']:
        r += 1
        write_section_header(ws6, r, '  HIGH VALUE TRANSACTIONS >= Rs 2 LAKH', 4)
        r += 1
        write_table_header(ws6, r, ['Date', 'Description', 'Type', 'Amount (Rs)'])
        r += 1
        for t in d_comp['high_value_txns']:
            clr = COLOR_GREEN if t.get('type') == 'CR' else COLOR_RED
            write_data_row(ws6, r, [t.get('date',''), t.get('desc','')[:80], t.get('type',''), f'Rs {t.get("amount",0):,.0f}'],
                           colors=[COLOR_MUTED, COLOR_WHITE, clr, clr])
            r += 1
    if d_comp['str_candidates']:
        r += 1
        write_section_header(ws6, r, '  STR CANDIDATES — SINGLE DAY >= Rs 5 LAKH', 3)
        r += 1
        write_table_header(ws6, r, ['Date', 'Daily Total (Rs)', 'Transactions'])
        r += 1
        for s in d_comp['str_candidates']:
            write_data_row(ws6, r, [s.get('date',''), f'Rs {s.get("total",0):,.0f}', f'{s.get("count",0)} txns'],
                           colors=[COLOR_MUTED, COLOR_YELLOW, COLOR_WHITE])
            r += 1
    set_col_widths(ws6, [28, 55, 16, 22])

    dark_fill = PatternFill('solid', fgColor=COLOR_DARK)
    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = COLOR_GOLD
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 1,
                                 min_col=1, max_col=max(ws.max_column or 1, 8)):
            for cell in row:
                if not cell.fill or cell.fill.patternType in (None, 'none'):
                    cell.fill = dark_fill
                if not cell.font or cell.font.color is None:
                    cell.font = Font(color=COLOR_WHITE, size=9, name='Calibri')
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 30
        for col_letter in ['C', 'D', 'E', 'F']:
            if ws.column_dimensions[col_letter].width < 12:
                ws.column_dimensions[col_letter].width = 14

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = secure_filename(filename.replace('.pdf', ''))
    return send_file(output,
                     download_name=f'AarogyamFin_Report_{safe_name}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/dashboard/export/pdf', methods=['GET'])
def dashboard_export_pdf():
    is_logged_in, user_email, user_id = _get_current_user()
    if not is_logged_in:
        return redirect('/login')

    cached_hash = session.get('file_hash')
    if not cached_hash:
        return "No data available. Please upload a PDF first.", 400
    cached = _cache_get(cached_hash)
    if not cached:
        return "Session expired. Please re-upload.", 400

    data     = run_dashboard(cached['transactions'])
    filename = session.get('file_name', 'statement')

    # ── TRACK: PDF report exported ──
    ph_track(user_id, event='pdf_report_exported', props={
        'file_name':         filename,
        'transaction_count': len(cached['transactions']),
        'export_type':       'pdf_13page',
        'total_cr':          data.get('total_cr', 0),
        'total_dr':          data.get('total_dr', 0),
    })

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, KeepTogether, PageBreak
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate
    from reportlab.pdfgen import canvas as pdfcanvas
    from datetime import date

    C_GOLD    = colors.HexColor('#C9A84C')
    C_GOLD_L  = colors.HexColor('#E8CC7A')
    C_BG      = colors.HexColor('#060810')
    C_SURFACE = colors.HexColor('#0D1117')
    C_SURFACE2= colors.HexColor('#111827')
    C_BORDER  = colors.HexColor('#1f2937')
    C_TEXT    = colors.HexColor('#E8E4D9')
    C_MUTED   = colors.HexColor('#8A8070')
    C_GREEN   = colors.HexColor('#4ade80')
    C_RED     = colors.HexColor('#f87171')
    C_YELLOW  = colors.HexColor('#fbbf24')
    C_BLUE    = colors.HexColor('#60a5fa')

    W = A4[0] - 40*mm

    def S(name, **kw):
        base = dict(fontName='Helvetica', fontSize=9, textColor=C_TEXT,
                    leading=14, spaceAfter=0, spaceBefore=0)
        base.update(kw)
        return ParagraphStyle(name, **base)

    s_cover_title  = S('ct',  fontSize=36, fontName='Helvetica-Bold', textColor=C_GOLD,   leading=40, alignment=TA_CENTER)
    s_cover_sub    = S('cs',  fontSize=13, textColor=C_TEXT,          leading=20, alignment=TA_CENTER)
    s_cover_meta   = S('cm',  fontSize=9,  textColor=C_MUTED,         leading=14, alignment=TA_CENTER)
    s_page_title   = S('pt',  fontSize=18, fontName='Helvetica-Bold', textColor=C_GOLD,   leading=22, spaceAfter=2)
    s_section      = S('sh',  fontSize=11, fontName='Helvetica-Bold', textColor=C_GOLD_L, leading=16, spaceBefore=10, spaceAfter=4)
    s_body         = S('b',   fontSize=8.5,textColor=C_TEXT,          leading=13)
    s_muted        = S('m',   fontSize=8,  textColor=C_MUTED,         leading=12)
    s_tag          = S('tg',  fontSize=7,  textColor=C_GOLD,          leading=10, fontName='Helvetica-Bold')
    s_footer       = S('ft',  fontSize=7,  textColor=C_MUTED,         leading=10, alignment=TA_CENTER)

    def tbl(rows, col_widths, style_extra=None):
        base_style = [
            ('BACKGROUND',   (0,0), (-1,-1), C_SURFACE),
            ('ROWBACKGROUNDS',(0,0),(-1,-1), [C_SURFACE, C_SURFACE2]),
            ('TEXTCOLOR',    (0,0), (-1,-1), C_TEXT),
            ('FONTNAME',     (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE',     (0,0), (-1,-1), 8.5),
            ('LEADING',      (0,0), (-1,-1), 13),
            ('GRID',         (0,0), (-1,-1), 0.3, C_BORDER),
            ('LEFTPADDING',  (0,0), (-1,-1), 7),
            ('RIGHTPADDING', (0,0), (-1,-1), 7),
            ('TOPPADDING',   (0,0), (-1,-1), 5),
            ('BOTTOMPADDING',(0,0), (-1,-1), 5),
            ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ]
        if style_extra:
            base_style.extend(style_extra)
        t = Table(rows, colWidths=col_widths)
        t.setStyle(TableStyle(base_style))
        return t

    def hdr_tbl(rows, col_widths, style_extra=None):
        s = [
            ('BACKGROUND',   (0,0), (-1, 0), colors.HexColor('#1a1500')),
            ('BACKGROUND',   (0,1), (-1,-1), C_SURFACE),
            ('ROWBACKGROUNDS',(0,1),(-1,-1), [C_SURFACE, C_SURFACE2]),
            ('TEXTCOLOR',    (0,0), (-1, 0), C_GOLD),
            ('TEXTCOLOR',    (0,1), (-1,-1), C_TEXT),
            ('FONTNAME',     (0,0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME',     (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE',     (0,0), (-1,-1), 8.5),
            ('LEADING',      (0,0), (-1,-1), 13),
            ('GRID',         (0,0), (-1,-1), 0.3, C_BORDER),
            ('LINEBELOW',    (0,0), (-1, 0), 0.8, C_GOLD),
            ('LEFTPADDING',  (0,0), (-1,-1), 7),
            ('RIGHTPADDING', (0,0), (-1,-1), 7),
            ('TOPPADDING',   (0,0), (-1,-1), 5),
            ('BOTTOMPADDING',(0,0), (-1,-1), 5),
            ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ]
        if style_extra:
            s.extend(style_extra)
        t = Table(rows, colWidths=col_widths)
        t.setStyle(TableStyle(s))
        return t

    def kv(label, value, val_color=None):
        vc = val_color or C_TEXT
        lp = Paragraph(label, S('kl', fontSize=8, textColor=C_MUTED, fontName='Helvetica'))
        vp = Paragraph(str(value), S('kv', fontSize=8.5, textColor=vc, fontName='Helvetica-Bold'))
        return [lp, vp]

    def divider():
        return HRFlowable(width='100%', thickness=0.4, color=C_BORDER, spaceAfter=6, spaceBefore=6)

    def section_header(text, tag=''):
        items = []
        if tag:
            items.append(Paragraph(tag, s_tag))
        items.append(Paragraph(text, s_page_title))
        items.append(divider())
        return items

    def alert(text, level='warn'):
        color = C_YELLOW if level=='warn' else C_RED if level=='danger' else C_GREEN
        bg    = colors.HexColor('#1a1200') if level=='warn' else colors.HexColor('#1a0000') if level=='danger' else colors.HexColor('#001a08')
        border= C_YELLOW if level=='warn' else C_RED if level=='danger' else C_GREEN
        p = Paragraph(text, S('al', fontSize=8.5, textColor=color, leading=13))
        t = Table([[p]], colWidths=[W])
        t.setStyle(TableStyle([
            ('BACKGROUND',  (0,0),(-1,-1), bg),
            ('LINEBEFORE',  (0,0),(0,-1), 3, border),
            ('LEFTPADDING', (0,0),(-1,-1), 10),
            ('RIGHTPADDING',(0,0),(-1,-1), 10),
            ('TOPPADDING',  (0,0),(-1,-1), 8),
            ('BOTTOMPADDING',(0,0),(-1,-1), 8),
        ]))
        return t

    def metrics_row(items):
        n = len(items)
        col_w = W / n
        row_vals  = [[Paragraph(str(v), S('mv2', fontSize=15, fontName='Helvetica-Bold',
                      textColor=c or C_GOLD, leading=19, alignment=TA_CENTER)) for _,v,c in items]]
        row_labels= [[Paragraph(l, S('ml2', fontSize=7, textColor=C_MUTED,
                      leading=10, alignment=TA_CENTER)) for l,_,__ in items]]
        combined = [row_vals[0], row_labels[0]]
        t = Table(combined, colWidths=[col_w]*n)
        t.setStyle(TableStyle([
            ('BACKGROUND',   (0,0),(-1,-1), C_SURFACE),
            ('GRID',         (0,0),(-1,-1), 0.3, C_BORDER),
            ('TOPPADDING',   (0,0),(-1,-1), 10),
            ('BOTTOMPADDING',(0,0),(-1,-1), 8),
            ('LEFTPADDING',  (0,0),(-1,-1), 4),
            ('RIGHTPADDING', (0,0),(-1,-1), 4),
        ]))
        return t

    today_str   = date.today().strftime('%d %B %Y')
    fname_clean = filename.replace('.pdf','').replace('_',' ')

    def draw_page(canv, doc):
        canv.saveState()
        canv.setFillColor(C_BG)
        canv.rect(0, 0, A4[0], A4[1], fill=1, stroke=0)
        canv.setFillColor(C_GOLD)
        canv.rect(0, A4[1]-2*mm, A4[0], 2*mm, fill=1, stroke=0)
        canv.setFillColor(C_MUTED)
        canv.setFont('Helvetica', 7)
        canv.drawString(20*mm, 12*mm, f'AarogyamFin Financial Report — {fname_clean}')
        canv.drawRightString(A4[0]-20*mm, 12*mm, f'Page {doc.page} | Generated {today_str}')
        canv.setStrokeColor(C_BORDER)
        canv.setLineWidth(0.3)
        canv.line(20*mm, 16*mm, A4[0]-20*mm, 16*mm)
        canv.restoreState()

    di   = data['itr']
    dl   = data['loan']
    dc   = data['compliance']
    da   = data['audit']
    de   = data['expense']
    dg   = data['gstr1']
    inc  = data.get('income', {})
    obl  = data.get('obligations', {})
    cf   = data.get('cashflow', {})
    rf   = data.get('red_flags', {})

    total_txns = data['total_txns']
    total_cr   = data['total_cr']
    total_dr   = data['total_dr']

    story = []

    # PAGE 1 — COVER
    story.append(Spacer(1, 40*mm))
    story.append(Paragraph('AarogyamFin', s_cover_title))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph('Financial Intelligence Report', s_cover_sub))
    story.append(Spacer(1, 8*mm))
    story.append(HRFlowable(width='60%', thickness=0.6, color=C_GOLD, hAlign='CENTER', spaceAfter=8, spaceBefore=0))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph(fname_clean, S('fn', fontSize=11, textColor=C_TEXT, leading=16, alignment=TA_CENTER)))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(f'Generated on {today_str}', s_cover_meta))
    story.append(Paragraph(f'Powered by AarogyamFin — aarogyamfin.com', s_cover_meta))
    story.append(Spacer(1, 16*mm))

    cover_data = [
        ['Total Transactions', f'{total_txns:,}',    'Total Credits',  f'₹{total_cr:,.0f}'],
        ['Total Debits',       f'₹{total_dr:,.0f}',  'Real Income',    f'₹{di["real_income_total"]:,.0f}'],
        ['Credit Profile',     dl['credit_indicator'],'FOIR',          f'{dl["foir"]}%'],
        ['Compliance Risk',    dc['risk_level'],       'Suggested ITR', di['suggested_itr'].split('(')[0].strip()],
    ]
    cover_tbl = Table(cover_data, colWidths=[45*mm, 45*mm, 45*mm, 45*mm])
    cover_tbl.setStyle(TableStyle([
        ('BACKGROUND',   (0,0),(-1,-1), C_SURFACE),
        ('ROWBACKGROUNDS',(0,0),(-1,-1),[C_SURFACE, C_SURFACE2]),
        ('TEXTCOLOR',    (0,0),(0,-1), C_MUTED),
        ('TEXTCOLOR',    (1,0),(1,-1), C_TEXT),
        ('TEXTCOLOR',    (2,0),(2,-1), C_MUTED),
        ('TEXTCOLOR',    (3,0),(3,-1), C_TEXT),
        ('FONTNAME',     (0,0),(-1,-1), 'Helvetica'),
        ('FONTNAME',     (1,0),(1,-1), 'Helvetica-Bold'),
        ('FONTNAME',     (3,0),(3,-1), 'Helvetica-Bold'),
        ('FONTSIZE',     (0,0),(-1,-1), 9),
        ('GRID',         (0,0),(-1,-1), 0.3, C_BORDER),
        ('LINEABOVE',    (0,0),(-1, 0), 1.5, C_GOLD),
        ('LINEBELOW',    (0,-1),(-1,-1), 1.5, C_GOLD),
        ('LEFTPADDING',  (0,0),(-1,-1), 10),
        ('RIGHTPADDING', (0,0),(-1,-1), 10),
        ('TOPPADDING',   (0,0),(-1,-1), 8),
        ('BOTTOMPADDING',(0,0),(-1,-1), 8),
    ]))
    story.append(cover_tbl)
    story.append(Spacer(1, 8*mm))
    story.append(Paragraph(
        'This report is AI-generated based on bank statement data. It is intended for '
        'preliminary analysis purposes only. Please verify all figures with original '
        'bank statements before making financial or legal decisions.',
        S('disc', fontSize=7.5, textColor=C_MUTED, leading=12, alignment=TA_CENTER)
    ))
    story.append(PageBreak())

    # Remaining pages follow the same pattern as original — full content preserved
    # PAGE 2 — Transaction Overview
    story += section_header('Transaction Overview', 'SECTION 01')
    story.append(metrics_row([
        ('Total Transactions', f'{total_txns:,}',          C_GOLD),
        ('Total Credits',      f'₹{total_cr:,.0f}',        C_GREEN),
        ('Total Debits',       f'₹{total_dr:,.0f}',        C_RED),
        ('Net Flow',           f'₹{total_cr - total_dr:,.0f}', C_GREEN if total_cr > total_dr else C_RED),
    ]))
    story.append(Spacer(1, 5*mm))
    md = dl.get('monthly_data', {})
    months_sorted = sorted(md.keys())
    if months_sorted:
        story.append(Paragraph('Monthly Credit vs Debit', s_section))
        mhdr = [['Month', 'Credits (₹)', 'Debits (₹)', 'Net (₹)', 'Closing Balance']]
        mrows = []
        for m in months_sorted:
            mv  = md[m]
            cr  = mv.get('credits', 0)
            dr  = mv.get('debits', 0)
            net = cr - dr
            bal = mv.get('min_bal', 0)
            mrows.append([m, f'₹{cr:,.0f}', f'₹{dr:,.0f}', f'₹{net:,.0f}', f'₹{bal:,.0f}'])
        story.append(hdr_tbl(mhdr + mrows, [35*mm, 37*mm, 37*mm, 35*mm, 36*mm]))
    story.append(PageBreak())

    # PAGE 3 — Income Analysis
    story += section_header('Income Analysis', 'SECTION 02')
    story.append(metrics_row([
        ('Real Income Total',  f'₹{di["real_income_total"]:,.0f}',   C_GREEN),
        ('Salary Income',      f'₹{di["salary_total"]:,.0f}',        C_TEXT),
        ('Freelance/Business', f'₹{di["freelance_total"]:,.0f}',     C_YELLOW),
        ('Interest Income',    f'₹{di["interest_total"]:,.0f}',      C_TEXT),
    ]))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph('Income Source Breakdown', s_section))
    inc_rows = [['Income Source', 'Amount (₹)', 'Classification']]
    for src, label, cls in [
        ('salary_total',         'Salary / Payroll',         'Taxable — Schedule S'),
        ('freelance_total',      'Freelance / Consulting',   'Taxable — PGBP'),
        ('marketplace_total',    'Marketplace / Ecommerce',  'Taxable — PGBP'),
        ('interest_total',       'Interest / FD',            'Taxable — Other Sources'),
        ('loan_disbursal_total', 'Loan Disbursals',          'Non-Income (Liability)'),
        ('family_transfer_total','Family Transfers',         'Non-Income (Transfer)'),
        ('self_transfer_total',  'Self Transfers',           'Non-Income (Internal)'),
    ]:
        val = di.get(src, 0)
        if val > 0:
            inc_rows.append([label, f'₹{val:,.0f}', cls])
    story.append(hdr_tbl(inc_rows, [70*mm, 50*mm, 60*mm]))
    story.append(Spacer(1, 4*mm))
    emp_names = inc.get('employer_names', [])
    if emp_names:
        story.append(Paragraph('Employers / Salary Sources Detected', s_section))
        story.append(Paragraph(', '.join(emp_names), s_body))
        story.append(Spacer(1, 3*mm))
    top_cr = inc.get('top_10_sources', [])
    if top_cr:
        story.append(Paragraph('Top Credit Sources', s_section))
        cr_hdr = [['Rank', 'Sender / Source', 'Total Amount (₹)', 'Transactions', 'Avg per Txn (₹)']]
        cr_rows = []
        for i, s in enumerate(top_cr[:10], 1):
            cr_rows.append([str(i), s['sender'], f'₹{s["total_amt"]:,.0f}', str(s['count']), f'₹{s["avg_amt"]:,.0f}'])
        story.append(hdr_tbl(cr_hdr + cr_rows, [15*mm, 65*mm, 45*mm, 30*mm, 25*mm]))
    story.append(PageBreak())

    # PAGE 4 — ITR
    story += section_header('ITR / Tax Filing Analysis', 'SECTION 03')
    story.append(alert(f'Suggested ITR Form: {di["suggested_itr"]}', 'ok'))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph('Tax Deductions Identified', s_section))
    ded_data = [
        kv('Section 80C (LIC / PPF / ELSS / EPF)',  f'₹{di["section_80c_total"]:,.0f}  (Max ₹1,50,000)',  C_YELLOW),
        kv('Section 80D (Health Insurance)',          f'₹{di["section_80d_total"]:,.0f}  (Max ₹25,000)',    C_YELLOW),
    ]
    story.append(tbl(ded_data, [90*mm, 90*mm]))
    story.append(Spacer(1, 4*mm))
    ded_80c = di.get('deductions', {}).get('80C', [])
    if ded_80c:
        story.append(Paragraph('Section 80C Transactions', s_section))
        r80c = [['Date', 'Description', 'Amount (₹)']]
        for t in ded_80c[:15]:
            r80c.append([t.get('date',''), t.get('desc','')[:60], f'₹{t.get("amount",0):,.0f}'])
        story.append(hdr_tbl(r80c, [30*mm, 110*mm, 40*mm]))
        story.append(Spacer(1, 3*mm))
    hvc = di.get('high_value_credits', [])
    if hvc:
        story.append(Paragraph(f'High Value Credits ≥ ₹1 Lakh ({len(hvc)} transactions)', s_section))
        hvc_r = [['Date', 'Description', 'Amount (₹)', 'Type']]
        for t in hvc[:12]:
            hvc_r.append([t.get('date',''), t.get('desc','')[:60],
                          f'₹{t.get("amount",0):,.0f}', t.get('credit_type','').replace('_',' ').title()])
        story.append(hdr_tbl(hvc_r, [30*mm, 80*mm, 40*mm, 30*mm], [
            ('TEXTCOLOR', (2,1),(2,-1), C_GREEN),
        ]))
    story.append(PageBreak())

    # PAGE 5 — Expenses
    story += section_header('Expense Categorization', 'SECTION 04')
    story.append(metrics_row([
        ('Total Debits',        f'₹{de["total_debits"]:,.0f}',      C_RED),
        ('Business Expenses',   f'₹{de["business_total"]:,.0f}',    C_BLUE),
        ('Personal Expenses',   f'₹{de["personal_total"]:,.0f}',    C_RED),
        ('GST Input Eligible',  f'₹{de["gst_eligible_total"]:,.0f}',C_GREEN),
    ]))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph('Category-wise Spend Breakdown', s_section))
    cat_hdr = [['Category', 'Amount (₹)', '% of Total']]
    cat_rows = []
    tot_dr = de['total_debits'] or 1
    for cat, amt in list(de.get('category_totals', {}).items())[:20]:
        pct = round(amt / tot_dr * 100, 1)
        cat_rows.append([cat, f'₹{amt:,.0f}', f'{pct}%'])
    story.append(hdr_tbl(cat_hdr + cat_rows, [80*mm, 60*mm, 40*mm]))
    story.append(Spacer(1, 4*mm))
    gst_el = de.get('gst_eligible', [])
    if gst_el:
        story.append(Paragraph(f'GST Input Credit Eligible Transactions ({len(gst_el)})', s_section))
        gst_r = [['Date', 'Description', 'Amount (₹)']]
        for t in gst_el[:12]:
            gst_r.append([t.get('date',''), t.get('desc','')[:70], f'₹{t.get("amount",0):,.2f}'])
        story.append(hdr_tbl(gst_r, [30*mm, 110*mm, 40*mm]))
    story.append(PageBreak())

    # PAGE 6 — Obligations
    story += section_header('Obligations & EMI Analysis', 'SECTION 05')
    story.append(metrics_row([
        ('Total EMI Outflow',   f'₹{obl.get("total_emi_outflow",0):,.0f}',  C_RED),
        ('Avg Monthly EMI',     f'₹{obl.get("avg_monthly_emi",0):,.0f}',    C_YELLOW),
        ('FOIR',                f'{dl["foir"]}%',                            C_RED if dl["foir"]>50 else C_YELLOW if dl["foir"]>35 else C_GREEN),
        ('EMI Capacity Left',   f'₹{dl["remaining_emi_capacity"]:,.0f}/mo', C_GREEN),
    ]))
    story.append(Spacer(1, 4*mm))
    if dl['foir'] > 50:
        story.append(alert(f'⚠ High FOIR: {dl["foir"]}% — EMI obligations exceed 50% of income. Loan approval risk HIGH.', 'danger'))
    elif dl['foir'] > 35:
        story.append(alert(f'⚠ Moderate FOIR: {dl["foir"]}% — Banks prefer FOIR ≤ 40% for best approval chances.', 'warn'))
    else:
        story.append(alert(f'✓ Healthy FOIR: {dl["foir"]}% — Good repayment capacity.', 'ok'))
    story.append(Spacer(1, 4*mm))
    emi_txns = obl.get('emi_txns', da.get('emis', []))
    if emi_txns:
        story.append(Paragraph(f'EMI / Loan Repayment Transactions ({len(emi_txns)})', s_section))
        emi_r = [['Date', 'Description', 'Amount (₹)']]
        for t in emi_txns[:15]:
            emi_r.append([t.get('date',''), t.get('desc','')[:70], f'₹{t.get("amount",0):,.2f}'])
        story.append(hdr_tbl(emi_r, [30*mm, 110*mm, 40*mm], [('TEXTCOLOR', (2,1),(2,-1), C_RED)]))
        story.append(Spacer(1, 3*mm))
    cc_pat = obl.get('cc_pattern', [])
    if cc_pat:
        story.append(Paragraph('Credit Card Payment Pattern', s_section))
        cc_r = [['Month', 'Total Paid (₹)', 'No. of Payments', 'Pattern']]
        for row in cc_pat:
            cc_r.append([row['month'], f'₹{row["total_paid"]:,.0f}', str(row['count']), row['pattern']])
        story.append(hdr_tbl(cc_r, [40*mm, 50*mm, 40*mm, 50*mm]))
    story.append(PageBreak())

    # PAGE 7 — Loan & Credit
    story += section_header('Loan & Credit Assessment', 'SECTION 06')
    credit_color = C_GREEN if dl['credit_color']=='green' else C_RED if dl['credit_color']=='red' else C_YELLOW
    story.append(metrics_row([
        ('Credit Profile',      dl['credit_indicator'],              credit_color),
        ('Avg Monthly Income',  f'₹{dl["avg_monthly_credit"]:,.0f}',C_GREEN),
        ('Avg Monthly Spend',   f'₹{dl["avg_monthly_debit"]:,.0f}', C_RED),
        ('Avg Balance',         f'₹{dl["avg_balance"]:,.0f}',       C_GOLD),
    ]))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph('Key Credit Metrics', s_section))
    credit_kv = [
        kv('FOIR (Fixed Obligation to Income Ratio)',  f'{dl["foir"]}%', C_RED if dl["foir"]>50 else C_YELLOW if dl["foir"]>35 else C_GREEN),
        kv('DSCR (Debt Service Coverage Ratio)',       f'{dl["dscr"]}x' if dl.get('dscr') else 'N/A', C_TEXT),
        kv('Minimum Balance',                          f'₹{dl["min_balance"]:,.0f}', C_TEXT),
        kv('Negative Balance Months',                  str(dl['negative_months']), C_RED if dl['negative_months']>0 else C_GREEN),
        kv('Remaining EMI Capacity',                   f'₹{dl["remaining_emi_capacity"]:,.0f}/month', C_GREEN),
        kv('Estimated Loan Eligible',                  f'₹{dl["loan_eligible"]:,.0f}', C_GREEN),
        kv('Months Analyzed',                          str(dl['months_analyzed']), C_TEXT),
    ]
    story.append(tbl(credit_kv, [95*mm, 85*mm]))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph('Monthly Cash Flow', s_section))
    if months_sorted:
        cf_hdr = [['Month', 'Credits (₹)', 'Debits (₹)', 'Net (₹)']]
        cf_rows = []
        for m in months_sorted:
            mv  = md[m]
            cr  = mv.get('credits', 0)
            dr  = mv.get('debits', 0)
            net = cr - dr
            cf_rows.append([m, f'₹{cr:,.0f}', f'₹{dr:,.0f}', f'{"+" if net>=0 else ""}₹{net:,.0f}'])
        story.append(hdr_tbl(cf_hdr + cf_rows, [40*mm, 50*mm, 50*mm, 40*mm], [
            ('TEXTCOLOR', (3,1),(3,-1), C_GREEN),
        ]))
    story.append(PageBreak())

    # PAGE 8 — Balance & Cash Flow
    story += section_header('Balance & Cash Flow Analysis', 'SECTION 07')
    story.append(metrics_row([
        ('ABB — 3 Month',  f'₹{cf.get("abb_3m",0):,.0f}',  C_GOLD),
        ('ABB — 6 Month',  f'₹{cf.get("abb_6m",0):,.0f}',  C_TEXT),
        ('ABB — 12 Month', f'₹{cf.get("abb_12m",0):,.0f}', C_TEXT),
        ('Avg Net Flow',   f'₹{cf.get("avg_net_flow",0):,.0f}', C_GREEN if cf.get("avg_net_flow",0)>0 else C_RED),
    ]))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph(f'Balance Trend: {cf.get("trend_direction","N/A")}', s_section))
    eom = cf.get('eom_trend', [])
    if eom:
        eom_hdr = [['Month', 'End-of-Month Balance (₹)']]
        eom_rows = [[e['month'], f'₹{e["balance"]:,.0f}'] for e in eom]
        story.append(hdr_tbl(eom_hdr + eom_rows, [80*mm, 100*mm]))
        story.append(Spacer(1, 3*mm))
    story.append(Paragraph('Seasonality Detection', s_section))
    high_m = cf.get('high_months', [])
    low_m  = cf.get('low_months', [])
    if high_m:
        story.append(alert(f'📈 High Income Months: {", ".join(high_m)} — 25%+ above average', 'ok'))
    if low_m:
        story.append(alert(f'📉 Low Income Months: {", ".join(low_m)} — 25%+ below average', 'warn'))
    if not high_m and not low_m:
        story.append(alert('✓ Stable Income Pattern — No major seasonal variation detected.', 'ok'))
    story.append(Spacer(1, 4*mm))
    mn = cf.get('monthly_net', [])
    if mn:
        story.append(Paragraph('Monthly Net Cash Flow', s_section))
        mn_hdr = [['Month', 'Credits (₹)', 'Debits (₹)', 'Net (₹)', 'Status']]
        mn_rows = []
        for row in mn:
            mn_rows.append([row['month'], f'₹{row["credits"]:,.0f}', f'₹{row["debits"]:,.0f}',
                            f'{"+" if row["surplus"] else ""}₹{row["net"]:,.0f}',
                            '✓ Surplus' if row['surplus'] else '✗ Deficit'])
        story.append(hdr_tbl(mn_hdr + mn_rows, [35*mm, 42*mm, 42*mm, 38*mm, 23*mm], [
            ('TEXTCOLOR', (3,1),(3,-1), C_GREEN),
        ]))
    story.append(PageBreak())

    # PAGE 9 — Audit
    story += section_header('Audit & Reconciliation', 'SECTION 08')
    story.append(metrics_row([
        ('Reconciliation Score', f'{da["reconciliation_score"]}/100',
         C_GREEN if da["reconciliation_score"]>=80 else C_YELLOW if da["reconciliation_score"]>=50 else C_RED),
        ('Balance Mismatches',   str(da['mismatch_count']),  C_RED if da['mismatch_count']>0 else C_GREEN),
        ('Bounced Txns',         str(da['bounce_count']),    C_RED if da['bounce_count']>0 else C_GREEN),
        ('Duplicate Suspects',   str(da['duplicate_count']), C_YELLOW if da['duplicate_count']>0 else C_GREEN),
    ]))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph('Audit Summary', s_section))
    audit_kv = [
        kv('Reconciliation Score',    f'{da["reconciliation_score"]}/100'),
        kv('Balance Mismatches',      str(da['mismatch_count'])),
        kv('Cheque Bounces',          str(da['bounce_count'])),
        kv('ECS Returns',             str(obl.get('ecs_count', 0))),
        kv('EMIs Detected',           str(da['emi_count'])),
        kv('Total EMI Outflow',       f'₹{da["total_emi_outflow"]:,.0f}'),
        kv('EMI-to-Income Ratio',     f'{da["emi_to_income_ratio"]}%'),
        kv('Duplicate Suspects',      str(da['duplicate_count'])),
    ]
    story.append(tbl(audit_kv, [95*mm, 85*mm]))
    story.append(Spacer(1, 4*mm))
    bm = da.get('balance_mismatches', [])
    if bm:
        story.append(alert(f'⚠ {len(bm)} balance continuity issue(s) found — may indicate data gaps or parsing issues.', 'warn'))
        story.append(Spacer(1, 3*mm))
        bm_hdr = [['Date', 'Description', 'Expected Balance', 'Actual Balance', 'Difference']]
        bm_rows = []
        for m in bm[:10]:
            bm_rows.append([m.get('date',''), m.get('desc','')[:45],
                            f'₹{m.get("expected_balance",0):,.2f}',
                            f'₹{m.get("balance",0):,.2f}',
                            f'₹{m.get("diff",0):,.2f}'])
        story.append(hdr_tbl(bm_hdr + bm_rows, [28*mm, 65*mm, 35*mm, 35*mm, 17*mm], [
            ('TEXTCOLOR', (4,1),(4,-1), C_RED),
        ]))
    else:
        story.append(alert('✓ All balances verified — No continuity issues found.', 'ok'))
    story.append(Spacer(1, 4*mm))
    bounced = da.get('bounced', [])
    if bounced:
        story.append(Paragraph(f'Bounced / Returned Transactions ({len(bounced)})', s_section))
        b_hdr = [['Date', 'Description', 'Amount (₹)']]
        b_rows = [[t.get('date',''), t.get('desc','')[:70], f'₹{t.get("amount",0):,.2f}'] for t in bounced[:10]]
        story.append(hdr_tbl(b_hdr + b_rows, [30*mm, 110*mm, 40*mm], [('TEXTCOLOR', (2,1),(2,-1), C_RED)]))
    story.append(PageBreak())

    # PAGE 10 — Red Flags
    story += section_header('Red Flags & Fraud Detection', 'SECTION 09')
    flag_color = C_GREEN if rf.get('flag_color')=='green' else C_RED if rf.get('flag_color')=='red' else C_YELLOW
    story.append(metrics_row([
        ('Flag Score',          str(rf.get('flag_score', 0)),         flag_color),
        ('Circular Txns',       str(rf.get('circular_count', 0)),     C_RED if rf.get('circular_count',0)>0 else C_GREEN),
        ('Gambling/Crypto',     str(rf.get('gambling_count', 0)),     C_RED if rf.get('gambling_count',0)>0 else C_GREEN),
        ('Duplicate Groups',    str(rf.get('duplicate_groups', 0)),   C_YELLOW if rf.get('duplicate_groups',0)>0 else C_GREEN),
    ]))
    story.append(Spacer(1, 4*mm))
    flag_level = rf.get('flag_level', 'Low')
    if flag_level == 'High':
        story.append(alert(f'🚨 HIGH RISK: {rf.get("total_flags",0)} red flags detected. Immediate review required.', 'danger'))
    elif flag_level == 'Medium':
        story.append(alert(f'⚠ MEDIUM RISK: {rf.get("total_flags",0)} flag(s) detected. Review recommended.', 'warn'))
    else:
        story.append(alert(f'✓ LOW RISK: {rf.get("total_flags",0)} flag(s). Statement appears clean.', 'ok'))
    story.append(Spacer(1, 4*mm))
    circ = rf.get('circular_txns', [])
    if circ:
        story.append(Paragraph(f'Circular Transactions ({len(circ)}) — Same amount credited & debited within 7 days', s_section))
        c_hdr = [['Credit Date', 'Credit Desc', 'Debit Date', 'Debit Desc', 'Amount', 'Gap']]
        c_rows = [[c['credit_date'], c['credit_desc'][:30], c['debit_date'],
                   c['debit_desc'][:30], f'₹{c["amount"]:,.0f}', f'{c["days_gap"]}d'] for c in circ[:8]]
        story.append(hdr_tbl(c_hdr + c_rows, [25*mm, 45*mm, 25*mm, 45*mm, 25*mm, 15*mm]))
        story.append(Spacer(1, 3*mm))
    wd = rf.get('window_dress', [])
    if wd:
        story.append(Paragraph(f'Window Dressing Alerts ({len(wd)}) — Large deposits near month-end', s_section))
        for w in wd[:5]:
            story.append(alert(
                f'Month-End Deposit: {w["deposit_date"]} — ₹{w["amount"]:,.0f} — '
                f'{w["withdrawal_count"]} matching withdrawal(s) within 8 days', 'warn'))
        story.append(Spacer(1, 3*mm))
    gamb = rf.get('gambling_txns', [])
    if gamb:
        story.append(Paragraph(f'Gambling / Crypto / Speculative ({len(gamb)}) — ₹{rf.get("gambling_total",0):,.0f} total', s_section))
        g_hdr = [['Date', 'Description', 'Type', 'Amount (₹)']]
        g_rows = [[t.get('date',''), t.get('desc','')[:65], t.get('type',''), f'₹{t.get("amount",0):,.0f}'] for t in gamb[:12]]
        story.append(hdr_tbl(g_hdr + g_rows, [28*mm, 90*mm, 15*mm, 47*mm]))
        story.append(Spacer(1, 3*mm))
    pen = rf.get('penalty_txns', [])
    if pen:
        story.append(Paragraph(f'Penalty / Legal Charges ({len(pen)}) — ₹{rf.get("penalty_total",0):,.0f}', s_section))
        p_hdr = [['Date', 'Description', 'Amount (₹)']]
        p_rows = [[t.get('date',''), t.get('desc','')[:75], f'₹{t.get("amount",0):,.0f}'] for t in pen[:10]]
        story.append(hdr_tbl(p_hdr + p_rows, [28*mm, 112*mm, 40*mm], [('TEXTCOLOR', (2,1),(2,-1), C_RED)]))
    story.append(PageBreak())

    # PAGE 11 — Compliance
    story += section_header('Compliance Reporting (PMLA / RBI)', 'SECTION 10')
    risk_color_map = {'green': C_GREEN, 'yellow': C_YELLOW, 'red': C_RED}
    risk_c = risk_color_map.get(dc.get('risk_color','green'), C_GREEN)
    story.append(metrics_row([
        ('Risk Level',          dc['risk_level'],               risk_c),
        ('Risk Score',          str(dc['risk_score']),          risk_c),
        ('High Value Txns',     str(dc['high_value_count']),    C_YELLOW if dc['high_value_count']>5 else C_GREEN),
        ('STR Candidates',      str(dc['str_count']),           C_RED if dc['str_count']>0 else C_GREEN),
    ]))
    story.append(Spacer(1, 4*mm))
    if dc['form_61a_required']:
        story.append(alert(f'🚨 Form 61A / SFT Filing REQUIRED — Annual cash total ₹{dc["annual_cash_total"]:,.0f} exceeds ₹10L threshold.', 'danger'))
    if dc['str_candidates']:
        story.append(alert(f'⚠ {dc["str_count"]} STR Candidate(s) — Single day credits ≥ ₹5L. May require filing under PMLA.', 'warn'))
    if dc['structured_suspects']:
        story.append(alert(f'⚠ {dc["structured_count"]} Structuring Suspect(s) — Transactions between ₹1.8L–₹2L detected.', 'warn'))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph('Compliance Metrics', s_section))
    comp_kv = [
        kv('Risk Level',                    dc['risk_level']),
        kv('High Value Transactions (≥₹2L)', str(dc['high_value_count'])),
        kv('Cash Transactions',             str(dc['cash_count'])),
        kv('STR Candidates (≥₹5L/day)',     str(dc['str_count'])),
        kv('Structuring Suspects',          str(dc['structured_count'])),
        kv('Daily Limit Breaches (>₹50K)',  str(len(dc['daily_breaches']))),
        kv('Annual Cash Total',             f'₹{dc["annual_cash_total"]:,.0f}'),
        kv('Form 61A / SFT Required',       'YES — File immediately' if dc['form_61a_required'] else 'Not Required'),
    ]
    story.append(tbl(comp_kv, [95*mm, 85*mm]))
    story.append(Spacer(1, 4*mm))
    hvt = dc.get('high_value_txns', [])
    if hvt:
        story.append(Paragraph(f'High Value Transactions ≥ ₹2 Lakh ({len(hvt)})', s_section))
        hv_hdr = [['Date', 'Description', 'Type', 'Amount (₹)']]
        hv_rows = [[t.get('date',''), t.get('desc','')[:65], t.get('type',''), f'₹{t.get("amount",0):,.0f}'] for t in hvt[:15]]
        story.append(hdr_tbl(hv_hdr + hv_rows, [28*mm, 90*mm, 15*mm, 47*mm], [('TEXTCOLOR', (2,1),(2,-1), C_GREEN)]))
    story.append(PageBreak())

    # PAGE 12 — Top Debits
    story += section_header('Top Debit Destinations & Spend Patterns', 'SECTION 11')
    top_db = obl.get('top_10_debits', [])
    if top_db:
        story.append(Paragraph('Top 10 Debit Destinations (by total amount)', s_section))
        db_hdr = [['Rank', 'Destination', 'Total Amount (₹)', 'Transactions', 'Avg per Txn (₹)']]
        db_rows = [[str(i+1), d['dest'], f'₹{d["total"]:,.0f}', str(d['count']), f'₹{d["avg"]:,.0f}'] for i, d in enumerate(top_db)]
        story.append(hdr_tbl(db_hdr + db_rows, [15*mm, 70*mm, 45*mm, 30*mm, 20*mm]))
        story.append(Spacer(1, 6*mm))
    story.append(Paragraph('GSTR-1 Sales Classification', s_section))
    story.append(metrics_row([
        ('B2B Supplies',   f'₹{dg["total_b2b"]:,.0f}',   C_GREEN),
        ('B2C Supplies',   f'₹{dg["total_b2c"]:,.0f}',   C_TEXT),
        ('Export Supplies',f'₹{dg["total_export"]:,.0f}', C_YELLOW),
        ('Est. GST @18%',  f'₹{dg["estimated_gst"]:,.0f}',C_RED),
    ]))
    story.append(Spacer(1, 4*mm))
    gst_status = 'Filing Required' if dg['total_taxable'] > 0 else 'Verify with CA'
    story.append(alert(
        f'GST Status: {gst_status} — Total Taxable: ₹{dg["total_taxable"]:,.0f} | '
        f'Est. Liability: ₹{dg["estimated_gst"]:,.0f} | Filing: {dg["filing_status"]}',
        'warn' if dg['total_taxable'] > 0 else 'ok'
    ))
    story.append(PageBreak())

    # PAGE 13 — Scorecard
    story += section_header('Financial Health Scorecard', 'SECTION 12 — SUMMARY')
    story.append(Paragraph(
        'This scorecard summarizes key financial health indicators derived from the bank statement analysis.',
        s_body
    ))
    story.append(Spacer(1, 5*mm))

    def score_row(label, value, status, color):
        return [label, value, Paragraph(status, S('sr', fontSize=8.5, textColor=color, fontName='Helvetica-Bold', leading=13))]

    scorecard = [
        ['Parameter', 'Value', 'Status'],
        score_row('Real Income (Monthly Avg)', f'₹{dl["avg_monthly_credit"]:,.0f}',
                  '✓ Good' if dl['avg_monthly_credit'] > 30000 else '⚠ Low',
                  C_GREEN if dl['avg_monthly_credit'] > 30000 else C_YELLOW),
        score_row('FOIR', f'{dl["foir"]}%',
                  '✓ Healthy' if dl["foir"]<=35 else '⚠ Moderate' if dl["foir"]<=50 else '✗ High Risk',
                  C_GREEN if dl["foir"]<=35 else C_YELLOW if dl["foir"]<=50 else C_RED),
        score_row('Loan Eligibility', f'₹{dl["loan_eligible"]:,.0f}',
                  '✓ Eligible' if dl['loan_eligible']>0 else '✗ Not Eligible',
                  C_GREEN if dl['loan_eligible']>0 else C_RED),
        score_row('Credit Profile', dl['credit_indicator'],
                  '✓ Strong' if dl['credit_color']=='green' else '⚠ Moderate' if dl['credit_color']=='yellow' else '✗ Weak',
                  C_GREEN if dl['credit_color']=='green' else C_YELLOW if dl['credit_color']=='yellow' else C_RED),
        score_row('Reconciliation Score', f'{da["reconciliation_score"]}/100',
                  '✓ Good' if da['reconciliation_score']>=80 else '⚠ Average' if da['reconciliation_score']>=50 else '✗ Poor',
                  C_GREEN if da['reconciliation_score']>=80 else C_YELLOW if da['reconciliation_score']>=50 else C_RED),
        score_row('Compliance Risk', dc['risk_level'],
                  '✓ Low' if dc['risk_level']=='Low' else '⚠ Medium' if dc['risk_level']=='Medium' else '✗ High',
                  C_GREEN if dc['risk_level']=='Low' else C_YELLOW if dc['risk_level']=='Medium' else C_RED),
        score_row('Cheque Bounces', str(da['bounce_count']),
                  '✓ Clean' if da['bounce_count']==0 else '✗ Found',
                  C_GREEN if da['bounce_count']==0 else C_RED),
        score_row('Red Flag Level', rf.get('flag_level','Low'),
                  '✓ Low' if rf.get('flag_level')=='Low' else '⚠ Medium' if rf.get('flag_level')=='Medium' else '✗ High',
                  C_GREEN if rf.get('flag_level')=='Low' else C_YELLOW if rf.get('flag_level')=='Medium' else C_RED),
        score_row('Form 61A Required', 'Yes' if dc['form_61a_required'] else 'No',
                  '✗ File Required' if dc['form_61a_required'] else '✓ Not Required',
                  C_RED if dc['form_61a_required'] else C_GREEN),
    ]

    sc_tbl = Table(scorecard, colWidths=[90*mm, 55*mm, 35*mm])
    sc_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(-1, 0), colors.HexColor('#1a1500')),
        ('BACKGROUND',    (0,1),(-1,-1), C_SURFACE),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[C_SURFACE, C_SURFACE2]),
        ('TEXTCOLOR',     (0,0),(-1, 0), C_GOLD),
        ('TEXTCOLOR',     (0,1),(-1,-1), C_TEXT),
        ('FONTNAME',      (0,0),(-1, 0), 'Helvetica-Bold'),
        ('FONTNAME',      (0,1),(-1,-1), 'Helvetica'),
        ('FONTSIZE',      (0,0),(-1,-1), 9),
        ('LEADING',       (0,0),(-1,-1), 14),
        ('GRID',          (0,0),(-1,-1), 0.3, C_BORDER),
        ('LINEABOVE',     (0,0),(-1, 0), 1.5, C_GOLD),
        ('LINEBELOW',     (0,-1),(-1,-1),1.5, C_GOLD),
        ('LINEBELOW',     (0, 0),(-1, 0), 0.8, C_GOLD),
        ('LEFTPADDING',   (0,0),(-1,-1), 10),
        ('RIGHTPADDING',  (0,0),(-1,-1), 10),
        ('TOPPADDING',    (0,0),(-1,-1), 8),
        ('BOTTOMPADDING', (0,0),(-1,-1), 8),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
    ]))
    story.append(sc_tbl)
    story.append(Spacer(1, 8*mm))
    story.append(HRFlowable(width='100%', thickness=0.4, color=C_BORDER))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        'DISCLAIMER: This report is generated by AarogyamFin AI based on bank statement data provided by the user. '
        'It is intended for preliminary financial analysis only. All figures must be verified against original bank '
        'statements, Form 26AS, and other financial documents before use in legal, tax, or credit applications. '
        'AarogyamFin does not assume liability for decisions made based on this report. '
        'For professional advice, consult a Chartered Accountant or financial advisor.',
        S('disc2', fontSize=7.5, textColor=C_MUTED, leading=12)
    ))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        f'Report generated by AarogyamFin — www.aarogyamfin.com | {today_str}',
        s_footer
    ))

    buffer = BytesIO()
    doc = BaseDocTemplate(
        buffer, pagesize=A4,
        rightMargin=20*mm, leftMargin=20*mm,
        topMargin=18*mm, bottomMargin=22*mm,
    )
    frame    = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='main')
    template = PageTemplate(id='main', frames=frame, onPage=draw_page)
    doc.addPageTemplates([template])
    doc.build(story)

    buffer.seek(0)
    safe_name = secure_filename(filename.replace('.pdf', ''))
    return send_file(buffer,
        download_name=f'AarogyamFin_Report_{safe_name}.pdf',
        as_attachment=True,
        mimetype='application/pdf')


@app.route('/gstr2b', methods=['GET', 'POST'])
def gstr2b_page():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('coming_soon.html', is_logged_in=is_logged_in, user_email=user_email, product='GSTR-2B Reconciliation')


@app.route('/gstr2b_disabled', methods=['POST'])
def gstr2b_analyze():
    ip = request.remote_addr
    if is_rate_limited(ip):
        abort(429)
    is_logged_in, user_email, user_id = _get_current_user()
    if not is_logged_in:
        return redirect('/login')
    from core.gstr2b_recon import run_gstr2b_recon
    error_message = ''
    data          = None
    gstr2b_path   = None
    pr_path       = None
    gstr2b_file = request.files.get('gstr2b_excel')
    if not gstr2b_file or not gstr2b_file.filename:
        error_message = 'GSTR-2B Excel file is required.'
        return render_template('gstr2b.html', data=None, error_message=error_message, is_logged_in=is_logged_in, user_email=user_email)
    safe_g = secure_filename(gstr2b_file.filename)
    if not safe_g.lower().endswith(('.xlsx', '.xls')):
        error_message = 'GSTR-2B file must be .xlsx or .xls'
        return render_template('gstr2b.html', data=None, error_message=error_message, is_logged_in=is_logged_in, user_email=user_email)
    gstr2b_path = os.path.join(app.config['UPLOAD_FOLDER'], 'gstr2b_' + safe_g)
    gstr2b_file.save(gstr2b_path)
    pr_file = request.files.get('pr_excel')
    if not pr_file or not pr_file.filename:
        error_message = 'Purchase Register Excel file is required.'
        if gstr2b_path and os.path.exists(gstr2b_path):
            os.remove(gstr2b_path)
        return render_template('gstr2b.html', data=None, error_message=error_message, is_logged_in=is_logged_in, user_email=user_email)
    safe_p = secure_filename(pr_file.filename)
    if not safe_p.lower().endswith(('.xlsx', '.xls')):
        error_message = 'Purchase Register file must be .xlsx or .xls'
        if gstr2b_path and os.path.exists(gstr2b_path):
            os.remove(gstr2b_path)
        return render_template('gstr2b.html', data=None, error_message=error_message, is_logged_in=is_logged_in, user_email=user_email)
    pr_path = os.path.join(app.config['UPLOAD_FOLDER'], 'pr_' + safe_p)
    pr_file.save(pr_path)
    try:
        data = run_gstr2b_recon(gstr2b_path, pr_path)
    except Exception as e:
        logger.exception("GSTR-2B recon error: %s", e)
        error_message = f'Analysis error: {e}'
    finally:
        for fp in [gstr2b_path, pr_path]:
            if fp and os.path.exists(fp):
                os.remove(fp)
    return render_template('gstr2b.html', data=data, error_message=error_message, is_logged_in=is_logged_in, user_email=user_email)


@app.route('/gstr3b', methods=['GET', 'POST'])
def gstr3b_page():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('coming_soon.html', is_logged_in=is_logged_in, user_email=user_email, product='GSTR-3B Computation')


@app.route('/gstr3b_disabled', methods=['POST'])
def gstr3b_analyze():
    ip = request.remote_addr
    if is_rate_limited(ip):
        abort(429)
    is_logged_in, user_email, user_id = _get_current_user()
    if not is_logged_in:
        return redirect('/login')
    from core.gstr3b import run_gstr3b
    error_message     = ''
    data              = None
    transactions      = []
    sales_filepath    = None
    purchase_filepath = None
    bank_pdf = request.files.get('bank_pdf')
    if bank_pdf and bank_pdf.filename:
        safe_name = secure_filename(bank_pdf.filename)
        if not safe_name.lower().endswith('.pdf'):
            error_message = 'Bank file must be a PDF.'
        elif not _is_valid_pdf(bank_pdf):
            error_message = 'Invalid bank PDF.'
        else:
            fhash       = _file_hash(bank_pdf)
            locked_hash = _get_locked_hash(user_id)
            if locked_hash and locked_hash != fhash:
                error_message = NEW_STATEMENT_ERROR
            else:
                cached = _cache_get(fhash)
                if cached:
                    transactions = cached['transactions']
                else:
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
                    bank_pdf.save(filepath)
                    try:
                        transactions = parse_transactions(filepath)
                        _cache_set(fhash, transactions, safe_name)
                    except Exception as e:
                        error_message = f'Bank PDF parse error: {e}'
                    finally:
                        if os.path.exists(filepath):
                            os.remove(filepath)
                if not error_message and transactions and not locked_hash:
                    _set_locked_hash(user_id, fhash)
                    session['file_hash'] = fhash
                    session['file_name'] = safe_name
    else:
        cached_hash = session.get('file_hash')
        if cached_hash:
            cached = _cache_get(cached_hash)
            if cached:
                transactions = cached['transactions']
    if error_message:
        return render_template('gstr3b.html', data=None, error_message=error_message, is_logged_in=is_logged_in, user_email=user_email)
    sales_file = request.files.get('sales_excel')
    if sales_file and sales_file.filename:
        safe_s = secure_filename(sales_file.filename)
        if safe_s.lower().endswith(('.xlsx', '.xls')):
            sales_filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'sales_' + safe_s)
            sales_file.save(sales_filepath)
    purchase_file = request.files.get('purchase_excel')
    if purchase_file and purchase_file.filename:
        safe_p = secure_filename(purchase_file.filename)
        if safe_p.lower().endswith(('.xlsx', '.xls')):
            purchase_filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'purchase_' + safe_p)
            purchase_file.save(purchase_filepath)
    try:
        data = run_gstr3b(transactions, sales_filepath, purchase_filepath)
    except Exception as e:
        logger.exception("GSTR-3B error: %s", e)
        error_message = f'Analysis error: {e}'
    finally:
        for fp in [sales_filepath, purchase_filepath]:
            if fp and os.path.exists(fp):
                os.remove(fp)
    return render_template('gstr3b.html', data=data, error_message=error_message, is_logged_in=is_logged_in, user_email=user_email)


@app.route('/gst-calendar')
def gst_calendar():
    is_logged_in, user_email, user_id = _get_current_user()
    from datetime import date
    today = date.today()
    return render_template('gst_calendar.html',
        is_logged_in = is_logged_in,
        user_email   = user_email,
        today        = today.isoformat(),
        current_month= today.month,
        current_year = today.year,
    )


@app.route('/gstr1', methods=['GET', 'POST'])
def gstr1_page():
    is_logged_in, user_email, user_id = _get_current_user()
    return render_template('coming_soon.html', is_logged_in=is_logged_in, user_email=user_email, product='GSTR-1 Auto-Fill')


@app.route('/gstr1_disabled', methods=['POST'])
def gstr1_analyze():
    ip = request.remote_addr
    if is_rate_limited(ip):
        abort(429)
    is_logged_in, user_email, user_id = _get_current_user()
    if not is_logged_in:
        return redirect('/login')
    error_message = ''
    data          = None
    transactions  = []
    bank_pdf = request.files.get('bank_pdf')
    if bank_pdf and bank_pdf.filename:
        safe_name = secure_filename(bank_pdf.filename)
        if not safe_name.lower().endswith('.pdf'):
            error_message = 'Bank file must be a PDF.'
        elif not _is_valid_pdf(bank_pdf):
            error_message = 'Invalid bank PDF.'
        else:
            fhash       = _file_hash(bank_pdf)
            locked_hash = _get_locked_hash(user_id)
            if locked_hash and locked_hash != fhash:
                error_message = NEW_STATEMENT_ERROR
            else:
                cached = _cache_get(fhash)
                if cached:
                    transactions = cached['transactions']
                else:
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
                    bank_pdf.save(filepath)
                    try:
                        transactions = parse_transactions(filepath)
                        _cache_set(fhash, transactions, safe_name)
                    except Exception as e:
                        error_message = f'Bank PDF parse error: {e}'
                    finally:
                        if os.path.exists(filepath):
                            os.remove(filepath)
                if not error_message and transactions and not locked_hash:
                    _set_locked_hash(user_id, fhash)
                    session['file_hash'] = fhash
                    session['file_name'] = safe_name
    else:
        cached_hash = session.get('file_hash')
        if cached_hash:
            cached = _cache_get(cached_hash)
            if cached:
                transactions = cached['transactions']
    if error_message:
        return render_template('gstr1.html', data=None, error_message=error_message, is_logged_in=is_logged_in, user_email=user_email)
    invoice_file = request.files.get('invoice_excel')
    if not invoice_file or not invoice_file.filename:
        error_message = 'Invoice Excel file required.'
        return render_template('gstr1.html', data=None, error_message=error_message, is_logged_in=is_logged_in, user_email=user_email)
    safe_inv = secure_filename(invoice_file.filename)
    if not safe_inv.lower().endswith(('.xlsx', '.xls')):
        error_message = 'Invoice file must be .xlsx or .xls'
        return render_template('gstr1.html', data=None, error_message=error_message, is_logged_in=is_logged_in, user_email=user_email)
    inv_filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_inv)
    invoice_file.save(inv_filepath)
    try:
        data = run_gstr1(inv_filepath, transactions)
    except Exception as e:
        logger.exception("GSTR-1 analysis error: %s", e)
        error_message = f'Analysis error: {e}'
    finally:
        if os.path.exists(inv_filepath):
            os.remove(inv_filepath)
    return render_template('gstr1.html', data=data, error_message=error_message, is_logged_in=is_logged_in, user_email=user_email)


# ═══════════════════════════════════════════════════════════
#  MOBILE API ROUTES
# ═══════════════════════════════════════════════════════════

@app.route('/api/parse', methods=['POST'])
def api_parse():
    ip = request.remote_addr
    if is_rate_limited(ip):
        return jsonify({'error': 'Too many requests'}), 429
    file = request.files.get('pdf_file')
    if not file or not file.filename:
        return jsonify({'error': 'No file uploaded'}), 400
    safe_name = secure_filename(file.filename)
    if not safe_name.lower().endswith('.pdf'):
        return jsonify({'error': 'Only PDF files accepted'}), 400
    if not _is_valid_pdf(file):
        return jsonify({'error': 'Invalid PDF'}), 400
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
    file.save(filepath)
    file_size_kb = round(os.path.getsize(filepath) / 1024, 1)
    t0 = time.time()
    try:
        transactions = parse_transactions(filepath)
        parse_time   = round(time.time() - t0, 1)
        logger.info("API parse: %s → %d txns in %.1fs", safe_name, len(transactions), parse_time)
        ph_track(ip, event='pdf_parse_success', props={
            'context':            'mobile_api',
            'file_name':          safe_name,
            'file_size_kb':       file_size_kb,
            'transaction_count':  len(transactions),
            'parse_time_seconds': parse_time,
        })
    except Exception as e:
        parse_time = round(time.time() - t0, 1)
        logger.exception("API parse error: %s", safe_name)
        ph_track(ip, event='pdf_parse_failed', props={
            'context':            'mobile_api',
            'file_name':          safe_name,
            'file_size_kb':       file_size_kb,
            'error_type':         type(e).__name__,
            'error_message':      str(e)[:200],
            'parse_time_seconds': parse_time,
        })
        return jsonify({'error': str(e)}), 500
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)
    if not transactions:
        return jsonify({'error': 'No transactions found'}), 400
    return jsonify({
        'transactions': transactions,
        'count':        len(transactions),
        'parse_time':   parse_time,
        'filename':     safe_name,
    }), 200


@app.route('/api/payment/create-order', methods=['POST'])
def api_create_order():
    ip = request.remote_addr
    if is_rate_limited(ip):
        return jsonify({'error': 'Too many requests'}), 429
    data         = request.get_json()
    firebase_uid = data.get('firebase_uid', '')
    email        = data.get('email', '')
    plan_name    = data.get('plan', 'Basic')
    if plan_name not in PLAN_CONFIG:
        plan_name = 'Basic'
    if not firebase_uid:
        return jsonify({'error': 'No user ID'}), 400
    plan   = PLAN_CONFIG[plan_name]
    amount = plan['price'] * 100
    try:
        order = rzp_client.order.create({
            'amount':          amount,
            'currency':        'INR',
            'payment_capture': 1,
        })
        supabase.table('payments').insert({
            'amount':            plan['price'],
            'status':            'pending',
            'razorpay_order_id': order['id'],
            'plan':              plan_name,
        }).execute()
        ph_track(firebase_uid, event='payment_order_created', props={
            'order_id':   order['id'],
            'amount_inr': plan['price'],
            'plan':       plan_name,
            'context':    'mobile',
        })
        logger.info("Mobile order created: %s plan=%s for %s", order['id'], plan_name, email)
        return jsonify({
            'order_id': order['id'],
            'amount':   amount,
            'currency': 'INR',
            'key_id':   RAZORPAY_KEY_ID,
            'email':    email,
        })
    except Exception as e:
        logger.exception("Mobile order creation failed: %s", e)
        return jsonify({'error': 'Order creation failed'}), 500


@app.route('/api/payment/verify', methods=['POST'])
def api_verify_payment():
    ip = request.remote_addr
    if is_rate_limited(ip):
        return jsonify({'error': 'Too many requests'}), 429
    data                = request.get_json()
    firebase_uid        = data.get('firebase_uid', '')
    razorpay_order_id   = data.get('razorpay_order_id', '')
    razorpay_payment_id = data.get('razorpay_payment_id', '')
    razorpay_signature  = data.get('razorpay_signature', '')
    if not firebase_uid:
        return jsonify({'error': 'No user ID'}), 400
    try:
        msg      = f"{razorpay_order_id}|{razorpay_payment_id}"
        expected = hmac.new(RAZORPAY_KEY_SECRET.encode(), msg.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(expected, razorpay_signature):
            return jsonify({'error': 'Payment verification failed'}), 400
    except Exception as e:
        return jsonify({'error': 'Verification error'}), 500
    try:
        plan_name = data.get('plan', 'Basic')
        if plan_name not in PLAN_CONFIG:
            plan_name = 'Basic'
        plan = PLAN_CONFIG[plan_name]
        from datetime import datetime, timezone, timedelta
        expires_at = datetime.now(timezone.utc) + timedelta(hours=plan['validity_hours'])
        result = supabase.table('payments').update({
            'status':              'paid',
            'razorpay_payment_id': razorpay_payment_id,
            'plan':                plan_name,
        }).eq('razorpay_order_id', razorpay_order_id).execute()
        payment_id = result.data[0]['id'] if result.data else None
        supabase.table('chat_sessions').insert({
            'firebase_uid':        firebase_uid,
            'payment_id':          payment_id,
            'is_active':           True,
            'db_ready':            False,
            'plan':                plan_name,
            'input_tokens_limit':  plan['input_tokens'],
            'output_tokens_limit': plan['output_tokens'],
            'input_tokens_used':   0,
            'output_tokens_used':  0,
            'expires_at':          expires_at.isoformat(),
        }).execute()
        ph_track(firebase_uid, event='payment_completed', props={
            'amount_inr': plan['price'],
            'plan':       plan_name,
            'payment_id': razorpay_payment_id,
            'context':    'mobile',
        })
        logger.info("Mobile payment verified: plan=%s for %s", plan_name, firebase_uid)
        return jsonify({'success': True})
    except Exception as e:
        logger.exception("Mobile payment verify DB error: %s", e)
        return jsonify({'error': 'Session creation failed'}), 500


@app.route('/api/payment/status', methods=['POST'])
def api_payment_status():
    data         = request.get_json()
    firebase_uid = data.get('firebase_uid', '')
    if not firebase_uid:
        return jsonify({'has_access': False}), 200
    try:
        from datetime import datetime, timezone
        result = supabase.table('chat_sessions').select('*').eq(
            'firebase_uid', firebase_uid
        ).eq('is_active', True).order('created_at', desc=True).limit(1).execute()
        if result.data:
            s            = result.data[0]
            input_used   = s.get('input_tokens_used', 0)
            output_used  = s.get('output_tokens_used', 0)
            input_limit  = s.get('input_tokens_limit', 25000)
            output_limit = s.get('output_tokens_limit', 5000)
            plan_name    = s.get('plan', 'Basic')
            expires_at   = s.get('expires_at')
            if expires_at:
                exp = datetime.fromisoformat(expires_at.replace('Z', '+00:00'))
                if exp < datetime.now(timezone.utc):
                    return jsonify({'has_access': False, 'plan': 'Free'})
            if input_used < input_limit and output_used < output_limit:
                saved_txns        = s.get('transactions') or []
                actual_session_id = str(s['id'])
                if saved_txns and not get_db(actual_session_id):
                    try:
                        build_index(actual_session_id, saved_txns)
                        logger.info("SQLite rebuilt for session %s", actual_session_id)
                    except Exception as e:
                        logger.exception("SQLite rebuild failed: %s", e)
                return jsonify({
                    'has_access':          True,
                    'plan':                plan_name,
                    'session_id':          actual_session_id,
                    'input_tokens_used':   input_used,
                    'output_tokens_used':  output_used,
                    'input_tokens_limit':  input_limit,
                    'output_tokens_limit': output_limit,
                    'messages':            s.get('messages') or [],
                    'pdf_loaded':          len(saved_txns) > 0,
                })
        return jsonify({'has_access': False, 'plan': 'Free'})
    except Exception as e:
        logger.exception("Payment status error: %s", e)
        return jsonify({'has_access': False}), 200


@app.route('/api/chat/upload', methods=['POST'])
def api_chat_upload():
    ip = request.remote_addr
    if is_rate_limited(ip):
        return jsonify({'error': 'Too many requests'}), 429
    file       = request.files.get('pdf_file')
    session_id = request.form.get('session_id', '')
    if not file or not file.filename:
        return jsonify({'error': 'No file uploaded'}), 400
    if not session_id:
        return jsonify({'error': 'No session_id provided'}), 400
    safe_name = secure_filename(file.filename)
    if not safe_name.lower().endswith('.pdf'):
        return jsonify({'error': 'Only PDF files accepted'}), 400
    if not _is_valid_pdf(file):
        return jsonify({'error': 'Invalid PDF'}), 400
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
    file.save(filepath)
    file_size_kb = round(os.path.getsize(filepath) / 1024, 1)
    t0 = time.time()
    try:
        transactions = parse_transactions(filepath)
        parse_time   = round(time.time() - t0, 1)
        ph_track(session_id, event='pdf_parse_success', props={
            'context':            'mobile_chat_upload',
            'file_name':          safe_name,
            'file_size_kb':       file_size_kb,
            'transaction_count':  len(transactions),
            'parse_time_seconds': parse_time,
        })
    except Exception as e:
        parse_time = round(time.time() - t0, 1)
        ph_track(session_id, event='pdf_parse_failed', props={
            'context':            'mobile_chat_upload',
            'file_name':          safe_name,
            'file_size_kb':       file_size_kb,
            'error_type':         type(e).__name__,
            'error_message':      str(e)[:200],
            'parse_time_seconds': parse_time,
        })
        return jsonify({'error': str(e)}), 500
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)
    if not transactions:
        return jsonify({'error': 'No transactions found'}), 400
    row_count = build_index(session_id, transactions)
    try:
        supabase.table('chat_sessions').update({
            'transactions': transactions,
            'db_ready':     True,
        }).eq('id', session_id).execute()
    except Exception as e:
        logger.exception("Failed to save transactions: %s", e)
    return jsonify({
        'success':           True,
        'session_id':        session_id,
        'transaction_count': row_count,
    }), 200


@app.route('/api/chat/message', methods=['POST'])
def api_chat_message():
    ip = request.remote_addr
    if is_rate_limited(ip):
        return jsonify({'error': 'Too many requests'}), 429
    data         = request.get_json()
    session_id   = data.get('session_id', '')
    user_message = (data.get('message') or '').strip()[:500]
    history      = data.get('history') or []
    if not session_id:
        return jsonify({'error': 'No session_id'}), 400
    if not user_message:
        return jsonify({'error': 'Empty message'}), 400
    if not get_db(session_id):
        return jsonify({'error': 'No statement loaded. Upload PDF first.'}), 400
    try:
        session_data = supabase.table('chat_sessions').select(
            'input_tokens_used, input_tokens_limit, output_tokens_used, output_tokens_limit, expires_at'
        ).eq('id', session_id).limit(1).execute()
        if session_data.data:
            s            = session_data.data[0]
            input_used   = s.get('input_tokens_used', 0)
            input_limit  = s.get('input_tokens_limit', 25000)
            output_used  = s.get('output_tokens_used', 0)
            output_limit = s.get('output_tokens_limit', 5000)
            from datetime import datetime, timezone
            expires_at = s.get('expires_at')
            if expires_at:
                exp = datetime.fromisoformat(expires_at.replace('Z', '+00:00'))
                if exp < datetime.now(timezone.utc):
                    return jsonify({'error': 'Session expired. Please purchase a new plan.'}), 403
            if input_used >= input_limit:
                return jsonify({'error': 'Token limit reached. Please purchase a new plan.'}), 403
            if output_used >= output_limit:
                return jsonify({'error': 'Token limit reached. Please purchase a new plan.'}), 403
    except Exception as e:
        logger.exception("Token limit check failed: %s", e)
    result      = ai_chat(session_id, user_message, history)
    tokens_used = result.get('tokens_used', 0)
    try:
        existing = supabase.table('chat_sessions').select(
            'input_tokens_used, output_tokens_used, messages'
        ).eq('id', session_id).limit(1).execute()
        if existing.data:
            s                 = existing.data[0]
            existing_messages = s.get('messages') or []
            existing_messages.append({'role': 'user',      'content': user_message})
            existing_messages.append({'role': 'assistant', 'content': result['reply']})
            existing_messages = existing_messages[-100:]
            supabase.table('chat_sessions').update({
                'input_tokens_used':  s.get('input_tokens_used', 0) + tokens_used,
                'output_tokens_used': s.get('output_tokens_used', 0) + result.get('output_tokens', 0),
                'messages':           existing_messages,
            }).eq('id', session_id).execute()
    except Exception as e:
        logger.exception("Token/message update failed: %s", e)
    ph_track(session_id, event='chat_message_sent', props={
        'context':       'mobile',
        'tokens_used':   tokens_used,
        'query_length':  len(user_message),
        'reply_length':  len(result.get('reply', '')),
    })
    return jsonify({'reply': result['reply'], 'tokens_used': tokens_used}), 200


# ═══════════════════════════════════════════════════════════
#  FINANCIAL CONSOLIDATOR
# ═══════════════════════════════════════════════════════════

@app.route('/consolidator', methods=['GET'])
def consolidator_page():
    is_logged_in, user_email, user_id = _get_current_user()
    if not is_logged_in:
        return redirect('/login')
    active = _get_active_chat_session(user_id)
    if not active:
        return redirect('/pay')
    return render_template('consolidator.html', data=None, is_logged_in=is_logged_in, user_email=user_email)


@app.route('/consolidator', methods=['POST'])
def consolidator_analyze():
    ip = request.remote_addr
    if is_rate_limited(ip):
        abort(429)
    is_logged_in, user_email, user_id = _get_current_user()
    if not is_logged_in:
        return redirect('/login')
    active = _get_active_chat_session(user_id)
    if not active:
        return redirect('/pay')
    from parsers.form26as import parse_26as
    from parsers.form_ais  import parse_ais
    from core.financial_consolidator import consolidate
    error_message   = ''
    data            = None
    transactions    = []
    data_26as       = None
    data_ais        = None
    bank_pdf = request.files.get('bank_pdf')
    if bank_pdf and bank_pdf.filename:
        safe_name = secure_filename(bank_pdf.filename)
        if safe_name.lower().endswith('.pdf') and _is_valid_pdf(bank_pdf):
            fhash  = _file_hash(bank_pdf)
            cached = _cache_get(fhash)
            if cached:
                transactions = cached['transactions']
            else:
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
                bank_pdf.save(filepath)
                try:
                    transactions = parse_transactions(filepath)
                    _cache_set(fhash, transactions, safe_name)
                    session['file_hash'] = fhash
                    session['file_name'] = safe_name
                except Exception as e:
                    error_message = f'Bank PDF parse error: {e}'
                finally:
                    if os.path.exists(filepath):
                        os.remove(filepath)
    else:
        cached_hash = session.get('file_hash')
        if cached_hash:
            cached = _cache_get(cached_hash)
            if cached:
                transactions = cached['transactions']
    if error_message:
        return render_template('consolidator.html', data=None, error_message=error_message, is_logged_in=is_logged_in, user_email=user_email)
    f26as = request.files.get('form_26as')
    if f26as and f26as.filename:
        safe_name = secure_filename(f26as.filename)
        if safe_name.lower().endswith('.pdf'):
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], '26as_' + safe_name)
            f26as.save(filepath)
            try:
                data_26as = parse_26as(filepath)
            except Exception as e:
                logger.exception("26AS parse error: %s", e)
                error_message += f' | 26AS error: {e}'
            finally:
                if os.path.exists(filepath):
                    os.remove(filepath)
    fais = request.files.get('form_ais')
    if fais and fais.filename:
        safe_name = secure_filename(fais.filename)
        if safe_name.lower().endswith('.pdf'):
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'ais_' + safe_name)
            fais.save(filepath)
            try:
                data_ais = parse_ais(filepath)
            except Exception as e:
                logger.exception("AIS parse error: %s", e)
                error_message += f' | AIS error: {e}'
            finally:
                if os.path.exists(filepath):
                    os.remove(filepath)
    opening_balance = 0.0
    ob_raw = request.form.get('opening_balance', '0').replace(',', '').strip()
    try:
        opening_balance = float(ob_raw) if ob_raw else 0.0
    except Exception:
        opening_balance = 0.0
    if not transactions and not data_26as and not data_ais:
        error_message = 'Please upload at least one document (Bank PDF or 26AS or AIS).'
        return render_template('consolidator.html', data=None, error_message=error_message, is_logged_in=is_logged_in, user_email=user_email)
    try:
        data = consolidate(bank_transactions=transactions, data_26as=data_26as, data_ais=data_ais, opening_balance=opening_balance)
    except Exception as e:
        logger.exception("Consolidation error: %s", e)
        error_message = f'Consolidation error: {e}'
    return render_template('consolidator.html', data=data, error_message=error_message, is_logged_in=is_logged_in, user_email=user_email)


# ═══════════════════════════════════════════════════════════
#  BLOG ROUTES
# ═══════════════════════════════════════════════════════════

import markdown
from pathlib import Path

BLOGS_DIR = Path('blogs')

@app.route('/blog')
def blog_index():
    posts = []
    for f in sorted(BLOGS_DIR.glob('*.md'), reverse=True):
        content = f.read_text(encoding='utf-8')
        lines = content.strip().split('\n')
        title = lines[0].replace('# ', '')
        excerpt = next((l for l in lines[1:] if l.strip() and not l.startswith('#')), '')[:160]
        posts.append({'slug': f.stem, 'title': title, 'excerpt': excerpt})
    is_logged_in, user_email, _ = _get_current_user()
    return render_template('blog_index.html', posts=posts, is_logged_in=is_logged_in, user_email=user_email)

@app.route('/blog/<slug>')
def blog_post(slug):
    filepath = BLOGS_DIR / f'{slug}.md'
    if not filepath.exists():
        abort(404)
    content = filepath.read_text(encoding='utf-8')
    title = content.strip().split('\n')[0].replace('# ', '')
    html_content = markdown.markdown(content, extensions=['tables', 'fenced_code'])
    is_logged_in, user_email, _ = _get_current_user()
    return render_template('blog_post.html', title=title, content=html_content, slug=slug, is_logged_in=is_logged_in, user_email=user_email)


# ═══════════════════════════════════════════════════════════
#  ERROR HANDLERS
# ═══════════════════════════════════════════════════════════

@app.errorhandler(429)
def too_many_requests(e):
    return "Too many requests. Please wait a minute and try again.", 429


@app.errorhandler(413)
def file_too_large(e):
    return "File too large. Maximum size is 10 MB.", 413


if __name__ == '__main__':
    debug_mode = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(debug=debug_mode)
    